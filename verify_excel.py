import openpyxl
import os

# Test file path
file_path = r'E:\UnityProgram\Game1\Docs_Sorting\Game1_文档整理_完整中文版_v3.xlsx'
print(f'File exists: {os.path.exists(file_path)}')
print(f'File size: {os.path.getsize(file_path)} bytes')
print()

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f'Sheet count: {len(wb.sheetnames)}')
    print(f'Sheet names: {wb.sheetnames}')
    print()
    
    # Check 文档索引 sheet
    if '文档索引' in wb.sheetnames:
        ws = wb['文档索引']
        print(f'文档索引 sheet - Rows: {ws.max_row}, Cols: {ws.max_column}')
        print('First 10 rows:')
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(5, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is not None:
                    row_data.append(str(val)[:40])
                else:
                    row_data.append('')
            print(f'  Row {row_idx}: {row_data}')
    else:
        print('文档索引 sheet NOT FOUND')
        # Print all sheets content
        for sname in wb.sheetnames:
            ws = wb[sname]
            print(f'=== {sname} ===')
            for r in range(1, min(4, ws.max_row + 1)):
                row_vals = [str(ws.cell(row=r, column=c).value)[:30] if ws.cell(row=r, column=c).value else '' for c in range(1, min(4, ws.max_column + 1))]
                print(f'  {row_vals}')
            print()
except Exception as e:
    print(f'Error: {e}')
    import traceback
    traceback.print_exc()