import openpyxl
import os

file_path = r'E:\\UnityProgram\\Game1\\Docs_Sorting\\Game1_文档整理_完整中文版_v3.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

# Get docs index sheet
ws = wb['\\u6587\\u6863\\u7d22\\u5f15']
print(f'Total rows in \\u6587\\u6863\\u7d22\\u5f15: {ws.max_row}')

# Collect all file names from Excel
excel_files = set()
for row in range(2, ws.max_row + 1):
    path_val = ws.cell(row=row, column=2).value
    name_val = ws.cell(row=row, column=3).value
    if name_val:
        excel_files.add(name_val)

print(f'Files listed in Excel: {len(excel_files)}')
print('File names:')
for f in sorted(excel_files):
    print(f'  {f}')
print()

# Get actual .md files from Docs directory
docs_dir = r'E:\\UnityProgram\\Game1\\Docs'
actual_files = set()
for root, dirs, files in os.walk(docs_dir):
    for f in files:
        if f.endswith('.md'):
            actual_files.add(f)

print(f'Actual .md files in Docs: {len(actual_files)}')

# Compare
missing_in_excel = actual_files - excel_files
extra_in_excel = excel_files - actual_files

print(f'\\nMissing in Excel: {len(missing_in_excel)}')
for f in sorted(missing_in_excel):
    print(f'  {f}')

print(f'\\nExtra in Excel: {len(extra_in_excel)}')
for f in sorted(extra_in_excel):
    print(f'  {f}')