import openpyxl
import os

file_path = r'E:\\UnityProgram\\Game1\\Docs_Sorting\\Game1_文档整理_完整中文版_v3.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

# Find the index sheet (first sheet)
index_sheet_name = wb.sheetnames[0]
print(f'First sheet name: {index_sheet_name}')
ws = wb[index_sheet_name]

print(f'Total rows: {ws.max_row}')

# Collect all file names from Excel
excel_files = []
for row in range(2, ws.max_row + 1):
    name_val = ws.cell(row=row, column=3).value
    if name_val:
        excel_files.append(name_val)

print(f'Files listed in Excel: {len(excel_files)}')
for f in sorted(excel_files):
    print(f'  {f}')
print()

# Get actual .md files from Docs directory
docs_dir = r'E:\\UnityProgram\\Game1\\Docs'
actual_files = []
for root, dirs, files in os.walk(docs_dir):
    for f in files:
        if f.endswith('.md'):
            actual_files.append(f)

print(f'Actual .md files in Docs: {len(actual_files)}')
for f in sorted(actual_files):
    print(f'  {f}')
print()

# Compare
actual_set = set(actual_files)
excel_set = set(excel_files)
missing = actual_set - excel_set
extra = excel_set - actual_set

print(f'Missing in Excel: {len(missing)}')
for f in sorted(missing):
    print(f'  {f}')

print(f'Extra in Excel: {len(extra)}')
for f in sorted(extra):
    print(f'  {f}')