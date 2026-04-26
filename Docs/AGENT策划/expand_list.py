# -*- coding: utf-8 -*-
"""
Game1 项目数据导出工具 - 扩展版
增强和扩展项目数据
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

# 打开已存在的工作簿
wb = openpyxl.load_workbook(r"E:\UnityProgram\Game1\Docs\AGENT策划\list.xlsx")

# 定义样式
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
expansion_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
legendary_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  # 淡紫色

def set_header(ws, headers, row=1):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

def set_row(ws, data, row, col_start=1, is_expansion=False, is_legendary=False):
    for i, value in enumerate(data, col_start):
        cell = ws.cell(row=row, column=i, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border
        if is_legendary:
            cell.fill = legendary_fill
        elif is_expansion:
            cell.fill = expansion_fill

def add_expansion_marker(ws, row, total_cols, note=""):
    cell = ws.cell(row=row, column=total_cols + 1, value=f"[拓展] {note}")
    cell.fill = expansion_fill
    cell.alignment = center_alignment

def auto_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

# ============================================================
# 新增工作表: 成就系统 (Achievements)
# ============================================================
ws_achievements = wb.create_sheet("成就系统")

ach_headers = ["ID", "名称", "类型", "条件", "奖励", "备注"]
set_header(ws_achievements, ach_headers)

achievements_data = [
    # 基础成就
    ("Core.Achievement.FirstStep", "迈出第一步", "Tutorial", "完成第一次旅行", "gold:+100", "教程成就"),
    ("Core.Achievement.FirstBlood", "初战告捷", "Combat", "赢得第一场战斗", "exp:+50", "战斗教程"),
    ("Core.Achievement.FirstTrade", "初次交易", "Trade", "完成第一次交易", "item:HealthPotion:3", "交易教程"),
    ("Core.Achievement.FirstVictory", "首战告捷", "Combat", "击败第一个BOSS", "gold:+500", "BOSS挑战"),
    # 战斗成就
    ("Core.Achievement.Kill100", "百人斩", "Combat", "累计击败100个敌人", "atk:+5", "战斗累计"),
    ("Core.Achievement.Kill1000", "千人斩", "Combat", "累计击败1000个敌人", "atk:+10", "战斗累计"),
    ("Core.Achievement.NoDamage", "完美闪避", "Combat", "无伤通关一次战斗", "crit_rate:+3%", "战斗技巧"),
    ("Core.Achievement.LowHpVictory", "绝境求生", "Combat", "HP低于10%时获胜", "hp:+20", "战斗技巧"),
    # 经济成就
    ("Core.Achievement.Gold1000", "小有资产", "Economy", "累计获得1000金币", "trade_discount:+5%", "积累成就"),
    ("Core.Achievement.Gold10000", "富甲一方", "Economy", "累计获得10000金币", "trade_discount:+10%", "积累成就"),
    ("Core.Achievement.Gold100000", "腰缠万贯", "Economy", "累计获得100000金币", "trade_discount:+15%", "积累成就"),
    # 探索成就
    ("Core.Achievement.Explorer", "探索者", "Exploration", "探索50个不同节点", "travel_speed:+5%", "探索成就"),
    ("Core.Achievement.TreasureHunter", "寻宝猎人", "Exploration", "开启10个宝箱", "rare_item_chance:+5%", "探索成就"),
    ("Core.Achievement.MapComplete", "地图探索者", "Exploration", "收集所有地图碎片", "reveal_all_map:true", "探索成就"),
    # 社交成就
    ("Core.Achievement.FriendMaker", "广交朋友", "Social", "与10个不同NPC对话", "charisma:+2", "社交成就"),
    ("Core.Achievement.Trader", "精明商人", "Social", "交易次数达100次", "trade_bonus:+5%", "交易成就"),
    # 技能成就
    ("Core.Achievement.SkillMaster", "技能大师", "Skill", "解锁全部技能", "all_skill_level:+1", "技能成就"),
    ("Core.Achievement.UltimateUser", "终极之力", "Skill", "首次使用终极技能", "ult_cd_reduce:-10%", "技能成就"),
    # 拓展成就
    ("Core.Achievement.PerfectGame", "完美通关", "Legendary", "无伤通关全部BOSS", "gold:+10000,GR_card:1", "拓展设计 - 传说成就"),
    ("Core.Achievement.Richest", "世界首富", "Economy", "累计获得1000000金币", "passive_gold:+10/秒", "拓展设计 - 经济成就"),
    ("Core.Achievement.Prestige10", "十次轮回", "Prestige", "完成10次轮回", "prestige_point:+50%", "拓展设计 - 轮回成就"),
    ("Core.Achievement.CardCollector", "卡牌收藏家", "Collection", "收集全部卡牌", "card_draw_rate:+10%", "拓展设计 - 收集成就"),
]

row = 2
for ach in achievements_data:
    is_expansion = "拓展设计" in str(ach)
    set_row(ws_achievements, ach[:5], row, is_expansion=is_expansion)
    if len(ach) > 5:
        add_expansion_marker(ws_achievements, row, 5, ach[5])
    row += 1

auto_width(ws_achievements)

# ============================================================
# 新增工作表: 坐骑系统 (Mounts)
# ============================================================
ws_mounts = wb.create_sheet("坐骑系统")

mount_headers = ["ID", "名称", "速度加成", "容量加成", "特殊效果", "解锁条件", "备注"]
set_header(ws_mounts, mount_headers)

mounts_data = [
    # 基础坐骑
    ("Core.Mount.None", "无", "0%", "0", "-", "初始", "默认状态"),
    ("Core.Mount.Horse", "马", "+15%", "+10", "-", "10级解锁", "基础坐骑"),
    ("Core.Mount.Camel", "骆驼", "+20%", "+20", "沙漠地形额外+10%", "20级解锁", "沙漠专用"),
    ("Core.Mount.Wagon", "马车", "+5%", "+50", "-", "30级解锁", "高容量低速度"),
    # 拓展坐骑
    ("Core.Mount.FastHorse", "快马", "+30%", "+5", "-", "25级解锁", "拓展设计 - 速度型"),
    ("Core.Mount.WarHorse", "战马", "+15%", "+15", "战斗加成+5%", "35级解锁", "拓展设计 - 战斗型"),
    ("Core.Mount.Ship", "商船", "+10%", "+100", "水路旅行+50%", "解锁港口后", "拓展设计 - 水路交通"),
    ("Core.Mount.Dragon", "幼龙", "+50%", "+30", "战斗加成+10%", "50级解锁", "拓展设计 - 传说坐骑"),
]

row = 2
for mount in mounts_data:
    is_expansion = "拓展设计" in str(mount)
    set_row(ws_mounts, mount[:6], row, is_expansion=is_expansion)
    if len(mount) > 6:
        add_expansion_marker(ws_mounts, row, 6, mount[6])
    row += 1

auto_width(ws_mounts)

# ============================================================
# 新增工作表: 状态效果 (StatusEffects)
# ============================================================
ws_status = wb.create_sheet("状态效果")

status_headers = ["ID", "名称", "类型", "持续时间", "效果", "叠加方式", "备注"]
set_header(ws_status, status_headers)

status_data = [
    # 战斗状态
    ("Core.Status.Poison", "中毒", "Debuff", "3回合", "每回合-5HP", "可叠加5层", "持续伤害"),
    ("Core.Status.Burn", "灼烧", "Debuff", "2回合", "每回合-3HP", "可叠加3层", "持续伤害"),
    ("Core.Status.Stun", "眩晕", "Debuff", "1回合", "无法行动", "不可叠加", "控制效果"),
    ("Core.Status.Slow", "减速", "Debuff", "2回合", "速度-30%", "不可叠加", "行动延后"),
    ("Core.Status.Weak", "虚弱", "Debuff", "2回合", "攻击-20%", "不可叠加", "输出降低"),
    # 增益状态
    ("Core.Status.Shield", "护盾", "Buff", "3回合", "吸收10伤害", "不可叠加", "伤害护盾"),
    ("Core.Status.Haste", "加速", "Buff", "2回合", "速度+30%", "不可叠加", "行动提前"),
    ("Core.Status.Power", "强化", "Buff", "2回合", "攻击+20%", "不可叠加", "输出提升"),
    ("Core.Status.Regeneration", "再生", "Buff", "3回合", "每回合+5HP", "不可叠加", "持续恢复"),
    ("Core.Status.Invincible", "无敌", "Buff", "1回合", "免疫所有伤害", "不可叠加", "免疫一切"),
    # 拓展状态
    ("Core.Status.Freeze", "冰冻", "Debuff", "1回合", "无法行动+易伤20%", "不可叠加", "拓展设计 - 强控"),
    ("Core.Status.Bleed", "流血", "Debuff", "4回合", "每回合-2HP", "可叠加10层", "拓展设计 - 持续伤害"),
    ("Core.Status.Fear", "恐惧", "Debuff", "2回合", "逃跑概率+30%", "不可叠加", "拓展设计 - 降低士气"),
    ("Core.Status.Blessing", "祝福", "Buff", "3回合", "全属性+10%", "不可叠加", "拓展设计 - 全面提升"),
    ("Core.Status.ManaShield", "魔法护盾", "Buff", "3回合", "吸收魔法伤害50%", "不可叠加", "拓展设计 - 魔抗护盾"),
]

row = 2
for status in status_data:
    is_expansion = "拓展设计" in str(status)
    set_row(ws_status, status[:6], row, is_expansion=is_expansion)
    if len(status) > 6:
        add_expansion_marker(ws_status, row, 6, status[6])
    row += 1

auto_width(ws_status)

# ============================================================
# 新增工作表: 稀有度配置 (RarityConfig)
# ============================================================
ws_rarity = wb.create_sheet("稀有度配置")

rarity_headers = ["稀有度", "等级", "颜色代码", "掉落概率", "保底次数", "卡片倍率", "备注"]
set_header(ws_rarity, rarity_headers)

rarity_data = [
    ("N (普通)", 1, "#FFFFFF", "40%", "-", "1.0x", "最常见"),
    ("R (稀有)", 2, "#00FF00", "30%", "10抽", "1.2x", "较常见"),
    ("SR (超稀有)", 3, "#0066FF", "15%", "10抽", "1.5x", "稀有"),
    ("SSR (超级稀有)", 4, "#FF00FF", "10%", "20抽", "1.8x", "很稀有"),
    ("UR (终极稀有)", 5, "#FF6600", "4%", "50抽", "2.0x", "极稀有"),
    ("GR (传说稀有)", 6, "#FFD700", "1%", "100抽", "3.0x", "最高稀有"),
    # 扩展稀有度
    ("LR (神话稀有)", 7, "#FF0000", "0.5%", "200抽", "5.0x", "拓展设计 - 神话级"),
]

row = 2
for rarity in rarity_data:
    is_expansion = "拓展设计" in str(rarity)
    set_row(ws_rarity, rarity[:6], row, is_expansion=is_expansion)
    if len(rarity) > 6:
        add_expansion_marker(ws_rarity, row, 6, rarity[6])
    row += 1

auto_width(ws_rarity)

# ============================================================
# 新增工作表: 属性类型 (AttributeTypes)
# ============================================================
ws_attr = wb.create_sheet("属性类型")

attr_headers = ["属性ID", "名称", "简称", "效果说明", "成长方式", "备注"]
set_header(ws_attr, attr_headers)

attr_data = [
    # 六维属性
    ("Core.Attribute.Vitality", "体力", "VIT", "最大HP，决定续航能力", "食物/休息/升级", "生命基础"),
    ("Core.Attribute.Attack", "攻击", "ATK", "伤害输出，影响战斗伤害", "武器/战斗/升级", "输出基础"),
    ("Core.Attribute.Defense", "防御", "DEF", "减伤承受，影响生存能力", "护甲/休息/升级", "防护基础"),
    ("Core.Attribute.Speed", "速度", "SPD", "行动顺序/旅行移动速度", "装备/事件/升级", "节奏基础"),
    ("Core.Attribute.Charisma", "魅力", "CHA", "NPC态度/交易价格折扣", "对话/事件/升级", "社交基础"),
    ("Core.Attribute.Wisdom", "智慧", "WIS", "暴击率/事件判断/技能效果", "书籍/事件/升级", "技巧基础"),
    # 战斗属性
    ("Core.Attribute.CritRate", "暴击率", "CRIT", "暴击触发概率", "装备/技能/升级", "输出提升"),
    ("Core.Attribute.CritDamage", "暴击伤害", "CRIT_DMG", "暴击时的伤害倍率", "装备/技能/升级", "爆发提升"),
    ("Core.Attribute.Dodge", "闪避率", "DODGE", "躲避攻击的概率", "装备/技能/升级", "生存提升"),
    ("Core.Attribute.Block", "格挡率", "BLOCK", "减少伤害的概率", "装备/技能/升级", "防护提升"),
    # 拓展属性
    ("Core.Attribute.Luck", "幸运", "LUCK", "暴击/掉落/事件触发", "完成成就/特殊事件", "拓展设计"),
    ("Core.Attribute.MagicPower", "魔法力", "MP", "魔法技能伤害", "法杖/魔法技能", "拓展设计"),
    ("Core.Attribute.MagicResist", "魔法抗性", "MR", "减少魔法伤害", "装备/抗性技能", "拓展设计"),
]

row = 2
for attr in attr_data:
    is_expansion = "拓展设计" in str(attr)
    set_row(ws_attr, attr[:5], row, is_expansion=is_expansion)
    if len(attr) > 5:
        add_expansion_marker(ws_attr, row, 5, attr[5])
    row += 1

auto_width(ws_attr)

# ============================================================
# 新增工作表: 事件稀有度 (EventRarity)
# ============================================================
ws_event_rarity = wb.create_sheet("事件稀有度")

event_rarity_headers = ["稀有度", "出现频率", "事件价值", "处理方式", "特殊效果", "备注"]
set_header(ws_event_rarity, event_rarity_headers)

event_rarity_data = [
    ("Normal (普通)", "70%", "低-中", "可批量处理", "无特殊", "高频率事件"),
    ("Rare (稀有)", "25%", "中-高", "单独关注", "可能掉落稀有物品", "中等频率"),
    ("Legendary (传奇)", "5%", "极高", "独特体验", "必定掉落GR卡或传说物品", "低频率高回报"),
    # 拓展稀有度
    ("Mythic (神话)", "1%", "极限", "重大抉择", "可能改变游戏进程", "拓展设计 - 极低频率"),
]

row = 2
for er in event_rarity_data:
    is_expansion = "拓展设计" in str(er)
    set_row(ws_event_rarity, er[:5], row, is_expansion=is_expansion)
    if len(er) > 5:
        add_expansion_marker(ws_event_rarity, row, 5, er[5])
    row += 1

auto_width(ws_event_rarity)

# ============================================================
# 新增工作表: 游戏数值公式 (Formulas)
# ============================================================
ws_formulas = wb.create_sheet("数值公式")

form_headers = ["公式ID", "名称", "公式", "说明", "备注"]
set_header(ws_formulas, form_headers)

formulas_data = [
    ("Core.Formula.Damage", "伤害计算", "max(1, 攻击方攻击力 - 防御方防御力 * 0.5)", "基础伤害公式", ""),
    ("Core.Formula.CritDamage", "暴击伤害", "攻击力 * 2.0", "暴击时伤害", ""),
    ("Core.Formula.FinalDamage", "最终伤害", "实际伤害 * (暴击? 2.0 : 1.0) * 技能倍率", "最终结算伤害", ""),
    ("Core.Formula.CritRate", "暴击率", "min(50%, 智慧 / 100 + 装备暴击率)", "最大50%上限", ""),
    ("Core.Formula.OfflineReward", "离线收益", "在线收益 * 时间系数 * 离线加成", "离线时间上限24小时", ""),
    ("Core.Formula.TimeCoeff", "时间系数", "1h内:1.0x, 1-6h:0.8x, 6-24h:0.5x, >24h:0.2x", "递减系数", ""),
    ("Core.Formula.PrestigePoint", "轮回点数", "(100 + 等级 * 50) * (1 + 加成)", "轮回重置收益", ""),
    ("Core.Formula.TravelSpeed", "旅行速度", "基础速度 * (1 + 速度加成) * (1 + 坐骑加成)", "综合移动速度", ""),
    ("Core.Formula.GoldToHealth", "金币换生命", "10 + 5*(等级-1) 金币恢复 (10 + 5*等级) HP", "商贾专属技能", ""),
    ("Core.Formula.TradeBonus", "交易加成", "基础交易价 * (1 - 魅力/200) * (1 + 商贾加成)", "最终交易价格", ""),
    ("Core.Formula.FoodConsume", "食物消耗", "基础消耗 * 距离 * 天气系数", "旅行食物计算", ""),
    # 拓展公式
    ("Core.Formula.CombatExp", "战斗经验", "敌人等级 * 10 * (1 + 学者加成) * (1 + 事件buff)", "战斗胜利获得", "拓展设计"),
    ("Core.Formula.TradeExp", "交易经验", "交易金额 * 0.1 * (1 + 魅力加成)", "交易完成获得", "拓展设计"),
    ("Core.Formula.GoldDrop", "金币掉落", "敌人基础金币 * (1 + 盗贼加成) * 随机(0.8~1.2)", "击败敌人掉落", "拓展设计"),
]

row = 2
for formula in formulas_data:
    is_expansion = "拓展设计" in str(formula)
    set_row(ws_formulas, formula[:4], row, is_expansion=is_expansion)
    if len(formula) > 4:
        add_expansion_marker(ws_formulas, row, 4, formula[4])
    row += 1

auto_width(ws_formulas)

# 保存文件
output_path = r"E:\UnityProgram\Game1\Docs\AGENT策划\list.xlsx"
wb.save(output_path)
print(f"Excel文件已更新: {output_path}")
print(f"共 {len(wb.sheetnames)} 个工作表:")
for name in wb.sheetnames:
    print(f"  - {name}")
