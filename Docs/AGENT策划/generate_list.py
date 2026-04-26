# -*- coding: utf-8 -*-
"""
Game1 项目数据导出工具
将XML配置数据导出为Excel表格
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

# 创建工作簿
wb = openpyxl.Workbook()

# 定义样式
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 扩展设计标记颜色
expansion_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 浅黄色

def set_header(ws, headers, row=1):
    """设置表头"""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

def set_row(ws, data, row, col_start=1, is_expansion=False):
    """设置数据行"""
    for i, value in enumerate(data, col_start):
        cell = ws.cell(row=row, column=i, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border
        if is_expansion:
            cell.fill = expansion_fill

def auto_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def add_expansion_marker(ws, row, total_cols, note=""):
    """为扩展行添加标记"""
    cell = ws.cell(row=row, column=total_cols + 1, value=f"[拓展] {note}")
    cell.fill = expansion_fill
    cell.alignment = center_alignment

# ============================================================
# 工作表1: 物品表 (Items)
# ============================================================
ws_items = wb.active
ws_items.title = "物品表"

items_headers = ["ID", "名称", "类型", "重量(kg)", "关键属性", "描述", "来源", "备注"]
set_header(ws_items, items_headers)

items_data = [
    # 现有物品
    ("Core.Item.GoldCoin", "金币", "Money", 0.0, "-", "游戏基础货币单位", "初始配置", ""),
    ("Core.Item.Bacon", "培根", "Food", 0.5, "热量1200kJ", "高热量食物，迅速恢复体力", "初始配置", ""),
    ("Core.Item.Cabbage", "卷心菜", "Food", 1.0, "热量400kJ", "普通蔬菜，轻微恢复体力", "初始配置", ""),
    ("Core.Item.DriedMeat", "肉干", "Food", 0.3, "热量800kJ", "便携高热量食物，适合旅行", "初始配置", ""),
    ("Core.Item.Fruit", "水果", "Food", 0.2, "热量200kJ", "新鲜水果，轻度恢复", "初始配置", ""),
    ("Core.Item.ShortBlade", "短刀", "Weapon", 2.0, "伤害+5", "基础近战武器，轻便易用", "初始配置", ""),
    ("Core.Item.LongSword", "长剑", "Weapon", 4.0, "伤害+12", "标准近战武器，伤害更高", "初始配置", ""),
    ("Core.Item.Bow", "弓", "Weapon", 1.5, "伤害+8", "远程武器，攻击距离远", "初始配置", ""),
    ("Core.Item.Staff", "法杖", "Weapon", 2.5, "伤害+6", "魔法武器，可使用技能", "初始配置", ""),
    ("Core.Item.LeatherArmor", "皮甲", "Armor", 5.0, "防御+3", "基础护甲，轻便灵活", "初始配置", ""),
    ("Core.Item.ChainMail", "锁子甲", "Armor", 10.0, "防御+8", "中级护甲，提供较好防护", "初始配置", ""),
    ("Core.Item.Wood", "木材", "Material", 3.0, "-", "基础建筑材料", "初始配置", ""),
    ("Core.Item.IronOre", "铁矿石", "Material", 5.0, "-", "用于冶炼金属的原料", "初始配置", ""),
    ("Core.Item.IronIngot", "铁锭", "Material", 4.0, "-", "精炼后的金属，可制作装备", "初始配置", ""),
    ("Core.Item.Coal", "煤炭", "Material", 2.0, "-", "燃料，用于熔炉和锻造", "初始配置", ""),
    ("Core.Item.Firewood", "木柴", "Material", 1.0, "-", "基础燃料，用于生火", "初始配置", ""),
    ("Core.Item.HealthPotion", "生命药剂", "Consumable", 0.5, "恢复20HP", "立即恢复生命值", "初始配置", ""),
    ("Core.Item.SpeedPotion", "速度药剂", "Consumable", 0.3, "速度+50%持续30秒", "短时间内提升移动速度", "初始配置", ""),
    ("Core.Item.AncientArtifact", "古代遗物", "QuestItem", 1.0, "-", "神秘的古代物品，任务关键道具", "初始配置", ""),
    ("Core.Item.MapFragment", "地图碎片", "QuestItem", 0.1, "-", "地图碎片，收集后可显示隐藏区域", "初始配置", ""),
    # 拓展物品
    ("Core.Item.Spear", "长矛", "Weapon", 3.0, "伤害+10", "长距离刺击武器", "拓展设计", "可克制骑兵类敌人"),
    ("Core.Item.Axe", "战斧", "Weapon", 5.0, "伤害+15", "重型劈砍武器，爆发力强", "拓展设计", "攻击速度较慢"),
    ("Core.Item.Dagger", "匕首", "Weapon", 0.8, "伤害+3,暴击+5%", "轻便近身武器，适合偷袭", "拓展设计", "刺客职业标配"),
    ("Core.Item.Crossbow", "弩", "Weapon", 3.5, "伤害+12", "精准远程武器", "拓展设计", "无需持续拉弦"),
    ("Core.Item.PlateArmor", "板甲", "Armor", 15.0, "防御+15", "重型全身护甲", "拓展设计", "速度惩罚-20%"),
    ("Core.Item.ClothRobe", "布甲", "Armor", 1.0, "防御+2,魔抗+5", "轻便魔法防护装备", "拓展设计", "法师职业标配"),
    ("Core.Item.Herb", "草药", "Material", 0.3, "-", "可用于炼制药剂", "拓展设计", "炼金术材料"),
    ("Core.Item.Thread", "线团", "Material", 0.2, "-", "用于制作和修复装备", "拓展设计", "裁缝材料"),
    ("Core.Item.Gem", "宝石", "Material", 0.5, "-", "稀有装饰和强化材料", "拓展设计", "可用于装备镶嵌"),
    ("Core.Item.ManaPotion", "魔法药剂", "Consumable", 0.4, "恢复30MP", "恢复魔法值", "拓展设计", "法师职业需求"),
    ("Core.Item.antidote", "解毒剂", "Consumable", 0.3, "解除中毒", "解除负面状态", "拓展设计", "野外生存必备"),
    ("Core.Item.Torch", "火把", "Consumable", 0.5, "照明60秒", "黑暗环境照明", "拓展设计", "洞穴探索道具"),
    ("Core.Item.Rope", "绳索", "QuestItem", 1.0, "-", "多用途道具", "拓展设计", "攀爬和陷阱解除"),
    ("Core.Item.OldKey", "古旧钥匙", "QuestItem", 0.2, "-", "开启特定宝箱", "拓展设计", "解谜要素"),
]

for i, item in enumerate(items_data, 2):
    is_expansion = item[6] == "拓展设计"
    set_row(ws_items, item[:6], i, is_expansion=is_expansion)
    if is_expansion:
        add_expansion_marker(ws_items, i, 6, item[7] if len(item) > 7 else "")

auto_width(ws_items)

# ============================================================
# 工作表2: 角色表 (Actors)
# ============================================================
ws_actors = wb.create_sheet("角色表")

actors_headers = ["ID", "名称", "阵营", "最大HP", "攻击", "防御", "速度", "金币奖励", "经验奖励", "是Boss", "交互类型", "备注"]
set_header(ws_actors, actors_headers)

actors_data = [
    # 现有角色
    ("Core.Actor.Player", "玩家", "Player", 20, 3, 5, 1.0, 0, 0, "否", "-", "玩家角色"),
    ("Core.Actor.Bandit", "匪徒", "Hostile", 20, 2, 3, 0.8, 15, 5, "否", "Combat", "基础敌人"),
    ("Core.Actor.BanditElite", "精英匪徒", "Hostile", 40, 4, 5, 0.9, 35, 12, "否", "Combat", "进阶敌人"),
    ("Core.Actor.WildBeast", "野兽", "Hostile", 15, 4, 2, 1.2, 10, 3, "否", "Combat", "高攻速敌人"),
    ("Core.Actor.Merchant", "行商", "Neutral", 10, 1, 8, 0.5, 25, 5, "否", "Trade", "可交易NPC"),
    ("Core.Actor.RoadGangLeader", "路寇首领", "Hostile", 100, 8, 10, 0.7, 100, 50, "是", "Combat", "BOSS级敌人"),
    ("Core.Actor.Villager", "村民", "Neutral", 20, 0, 1, 0.5, 0, 0, "否", "Story", "剧情NPC"),
    ("Core.Actor.Guard", "守卫", "Authority", 50, 5, 8, 0.8, 0, 0, "否", "Combat", "城镇守卫"),
    ("Core.Actor.Scholar", "学者", "Neutral", 15, 0, 2, 0.6, 0, 0, "否", "Trade", "知识型NPC"),
    ("Core.Actor.Healer", "医者", "Neutral", 18, 0, 3, 0.5, 0, 0, "否", "Trade", "治疗型NPC"),
    # 拓展角色
    ("Core.Actor.Wolf", "野狼", "Hostile", 25, 5, 2, 1.3, 12, 4, "否", "Combat", "拓展设计 - 群居动物"),
    ("Core.Actor.Bear", "熊", "Hostile", 60, 10, 6, 0.6, 30, 15, "否", "Combat", "拓展设计 - 高防高攻"),
    ("Core.Actor.Goblin", "哥布林", "Hostile", 12, 3, 1, 1.0, 8, 2, "否", "Combat", "拓展设计 - 群居小怪"),
    ("Core.Actor.GoblinShaman", "哥布林萨满", "Hostile", 20, 4, 2, 0.8, 20, 8, "否", "Combat", "拓展设计 - 会施法"),
    ("Core.Actor.BanditChief", "山贼头目", "Hostile", 80, 7, 8, 0.75, 80, 40, "是", "Combat", "拓展设计 - 中期BOSS"),
    ("Core.Actor.Dragon", "幼龙", "Hostile", 150, 12, 12, 0.5, 200, 100, "是", "Combat", "拓展设计 - 终局BOSS"),
    ("Core.Actor.Blacksmith", "铁匠", "Friendly", 30, 2, 10, 0.4, 0, 0, "否", "Trade", "拓展设计 - 装备强化"),
    ("Core.Actor.Innkeeper", "旅店老板", "Friendly", 25, 1, 5, 0.5, 0, 0, "否", "Story", "拓展设计 - 休息恢复"),
    ("Core.Actor.GuildMaster", "公会会长", "Authority", 40, 4, 8, 0.6, 0, 0, "否", "Story", "拓展设计 - 任务发布"),
    ("Core.Actor.Thief", "盗贼", "Neutral", 18, 6, 3, 1.4, 20, 6, "否", "Combat", "拓展设计 - 高闪避"),
]

for i, actor in enumerate(actors_data, 2):
    is_expansion = "拓展设计" in str(actor)
    set_row(ws_actors, actor[:11], i, is_expansion=is_expansion)
    if len(actor) > 11:
        add_expansion_marker(ws_actors, i, 11, actor[11])

auto_width(ws_actors)

# ============================================================
# 工作表3: 技能表 (Skills)
# ============================================================
ws_skills = wb.create_sheet("技能表")

skills_headers = ["ID", "名称", "类型", "最大等级", "效果类型", "基础效果值", "每级增量", "冷却(秒)", "基础消耗", "触发条件", "备注"]
set_header(ws_skills, skills_headers)

skills_data = [
    # 现有技能
    ("Core.Skill.Passive.Merchant", "商贾本能", "Passive", 5, "trade_bonus", "6%", "6%/级", "-", "-", "None", "交易加成被动"),
    ("Core.Skill.Passive.Guardian", "守护之心", "Passive", 5, "defense_boost", "+2", "+1/级", "-", "-", "None", "防御加成被动"),
    ("Core.Skill.Passive.Scholar", "学者之心", "Passive", 5, "exp_bonus", "5%", "3%/级", "-", "-", "None", "经验加成被动"),
    ("Core.Skill.Passive.Healer", "医者之心", "Passive", 5, "heal_bonus", "8%", "4%/级", "-", "-", "None", "治疗加成被动"),
    ("Core.Skill.Passive.SwiftStep", "快捷步伐", "Passive", 3, "travel_speed", "4%", "4%/级", "-", "-", "InTravel", "旅行速度被动"),
    ("Core.Skill.Passive.CritMaster", "暴击精通", "Passive", 5, "crit_rate", "2%", "1%/级", "-", "-", "InCombat", "暴击率被动"),
    ("Core.Skill.Passive.Tenacity", "坚韧", "Passive", 5, "hp_boost", "5%", "3%/级", "-", "-", "None", "生命值加成被动"),
    ("Core.Skill.Active.Sweep", "横扫", "Active", 3, "area_damage", "50%", "10%/级", 30, 0, "InCombat", "范围伤害技能"),
    ("Core.Skill.Active.Heal", "治疗术", "Active", 3, "heal", "30%", "5%/级", 45, 0, "InCombat", "单体治疗技能"),
    ("Core.Skill.Active.Shield", "护盾", "Active", 3, "shield", "20%", "10%/级", 60, 0, "InCombat", "护盾吸收伤害"),
    ("Core.Skill.Active.GoldToHealth", "金币换生命", "Active", 3, "gold_to_health", "10HP/100金", "+5HP/级", 90, 100, "OutOfCombat", "消耗金币恢复HP"),
    ("Core.Skill.Active.SpeedBurst", "速度爆发", "Active", 5, "travel_speed_boost", "10%", "5%/级", 120, 0, "InTravel", "临时旅行加速"),
    ("Core.Skill.Ultimate.DeadlyStrike", "致命一击", "Ultimate", 1, "single_target_crit", "300%", "-", 120, 0, "InCombat", "单体高伤害终极技"),
    ("Core.Skill.Ultimate.GroupHeal", "全体治疗", "Ultimate", 1, "group_heal", "50%", "-", 180, 0, "InCombat", "群体治疗终极技"),
    # 拓展技能
    ("Core.Skill.Passive.Vitality", "活力", "Passive", 5, "hp_regen", "1HP/秒", "0.5HP/级", "-", "-", "None", "拓展设计 - 持续回血"),
    ("Core.Skill.Passive.Thief", "窃贼本能", "Passive", 5, "gold_bonus", "3%", "2%/级", "-", "-", "None", "拓展设计 - 金币获取"),
    ("Core.Skill.Passive.EagleEye", "鹰眼", "Passive", 3, "crit_damage", "10%", "5%/级", "-", "-", "InCombat", "拓展设计 - 暴击伤害"),
    ("Core.Skill.Active.DoubleStrike", "双击", "Active", 3, "double_attack", "1次", "-", 25, 0, "InCombat", "拓展设计 - 攻击两次"),
    ("Core.Skill.Active.Intimidate", "威慑", "Active", 3, "enemy_def_down", "-10%", "-5%/级", 40, 0, "InCombat", "拓展设计 - 降低敌人防御"),
    ("Core.Skill.Active.Counter", "反击", "Active", 3, "counter_attack", "30%", "+10%/级", 50, 0, "InCombat", "拓展设计 - 受击时反击"),
    ("Core.Skill.Ultimate.TimeStop", "时间停止", "Ultimate", 1, "stun_all", "3秒", "-", 300, 0, "InCombat", "拓展设计 - 全屏控制"),
    ("Core.Skill.Ultimate.BloodPact", "鲜血契约", "Ultimate", 1, "sacrifice_hp", "消耗30%HP", "-", 200, 0, "InCombat", "拓展设计 - 消耗生命换取力量"),
]

for i, skill in enumerate(skills_data, 2):
    is_expansion = "拓展设计" in str(skill)
    set_row(ws_skills, skill[:10], i, is_expansion=is_expansion)
    if len(skill) > 10:
        add_expansion_marker(ws_skills, i, 10, skill[10])

auto_width(ws_skills)

# ============================================================
# 工作表4: 卡牌表 (Cards)
# ============================================================
ws_cards = wb.create_sheet("卡牌表")

cards_headers = ["ID", "名称", "类型", "稀有度", "属性倍率", "词缀", "售价", "备注"]
set_header(ws_cards, cards_headers)

cards_data = [
    # 现有卡牌 - 角色卡
    ("Core.Card.Character.Warrior", "战士", "Character", "R", "1.2x", "-", 100, "基础角色卡"),
    ("Core.Card.Character.Mage", "法师", "Character", "SR", "1.5x", "-", 500, "魔法输出"),
    ("Core.Card.Character.Healer", "治疗师", "Character", "SR", "1.4x", "-", 450, "治疗辅助"),
    ("Core.Card.Character.Rogue", "盗贼", "Character", "R", "1.3x", "-", 150, "高暴击"),
    ("Core.Card.Character.Berserker", "狂战士", "Character", "SSR", "1.8x", "-", 2000, "高攻低防"),
    # 现有卡牌 - 装备卡
    ("Core.Card.Equipment.Sword", "长剑", "Equipment", "N", "1.0x", "ATK_BONUS_10", 20, "基础武器"),
    ("Core.Card.Equipment.Shield", "盾牌", "Equipment", "N", "1.0x", "DEF_BONUS_10", 20, "基础防御"),
    ("Core.Card.Equipment.Staff", "法杖", "Equipment", "R", "1.2x", "MP_BONUS_50", 120, "魔法武器"),
    ("Core.Card.Equipment.Armor", "护甲", "Equipment", "R", "1.2x", "DEF_BONUS_25", 150, "防御装备"),
    ("Core.Card.Equipment.LegendaryBlade", "传奇之刃", "Equipment", "UR", "2.0x", "ATK_BONUS_100,CRIT_15", 10000, "顶级武器"),
    # 现有卡牌 - 技能卡
    ("Core.Card.Skill.Fireball", "火球术", "Skill", "R", "1.3x", "SKILL_DAMAGE_50", 180, "单体火系魔法"),
    ("Core.Card.Skill.Heal", "治疗术", "Skill", "R", "1.2x", "SKILL_HEAL_30", 160, "治疗魔法"),
    ("Core.Card.Skill.Thunder", "雷电术", "Skill", "SR", "1.5x", "SKILL_DAMAGE_80,STUN", 800, "带眩晕效果"),
    ("Core.Card.Skill.IceWall", "冰墙术", "Skill", "SR", "1.4x", "SKILL_DEF_60", 600, "防御型魔法"),
    # 现有卡牌 - 道具卡
    ("Core.Card.Item.HealthPotion", "生命药剂", "Item", "N", "1.0x", "USE_HEAL_100", 10, "恢复道具"),
    ("Core.Card.Item.ManaPotion", "魔法药剂", "Item", "N", "1.0x", "USE_MP_50", 10, "魔恢复道具"),
    ("Core.Card.Item.Elixir", "万能药", "Item", "R", "1.5x", "USE_BUFF_ALL", 300, "全属性提升"),
    # 现有卡牌 - 事件卡
    ("Core.Card.Event.Treasure", "宝箱事件", "Event", "R", "1.0x", "EVENT_GOLD_500", 80, "随机金币"),
    ("Core.Card.Event.Merchant", "商人事件", "Event", "SR", "1.3x", "EVENT_DISCOUNT_20", 500, "折扣机会"),
    ("Core.Card.Event.Dragon", "巨龙事件", "Event", "SSR", "2.0x", "EVENT_BOSS_SPAWN", 5000, "BOSS遭遇"),
    # GR级传说卡
    ("Core.Card.Legend.Destiny", "命运之人", "Character", "GR", "3.0x", "ATK+200,DEF+200,CRIT_30,EXP_50", 50000, "全能角色"),
    ("Core.Card.Legend.AncientPower", "远古之力", "Skill", "GR", "3.0x", "ULT_DAMAGE_500,CD_REDUCE_50", 80000, "究极技能"),
    ("Core.Card.Legend.Blessing", "神之祝福", "Item", "GR", "3.0x", "FULL_RECOVERY,REVIVE", 100000, "完美生存"),
    # 拓展卡牌
    ("Core.Card.Character.Paladin", "圣武士", "Character", "SR", "1.5x", "DEF_BONUS_30,HEAL_20", 600, "拓展设计 - 坦克治疗"),
    ("Core.Card.Character.Archer", "弓箭手", "Character", "R", "1.3x", "RANGE_ATTACK,CRIT_10", 180, "拓展设计 - 远程输出"),
    ("Core.Card.Equipment.Bow", "精灵长弓", "Equipment", "SR", "1.4x", "RANGE_DAMAGE_80", 400, "拓展设计 - 稀有武器"),
    ("Core.Card.Equipment.Ring", "力量戒指", "Equipment", "R", "1.2x", "ATK_BONUS_15", 200, "拓展设计 - 饰品"),
    ("Core.Card.Skill.Poison", "毒系魔法", "Skill", "R", "1.2x", "POISON_DAMAGE,DURATION", 150, "拓展设计 - 持续伤害"),
    ("Core.Card.Skill.Teleport", "传送术", "Skill", "SR", "1.3x", "INSTANT_MOVE,COOLDOWN", 350, "拓展设计 - 位移技能"),
    ("Core.Card.Event.Storm", "暴风雨", "Event", "R", "1.0x", "EVENT_GOLD_200,HP_LOSS", 100, "拓展设计 - 风险事件"),
    ("Core.Card.Event.Shrine", "古老神殿", "Event", "SR", "1.3x", "EVENT_BUFF_RANDOM", 450, "拓展设计 - 增益事件"),
]

for i, card in enumerate(cards_data, 2):
    is_expansion = "拓展设计" in str(card)
    set_row(ws_cards, card[:7], i, is_expansion=is_expansion)
    if len(card) > 7:
        add_expansion_marker(ws_cards, i, 7, card[7])

auto_width(ws_cards)

# ============================================================
# 工作表5: 事件表 (Events)
# ============================================================
ws_events = wb.create_sheet("事件表")

events_headers = ["ID", "名称", "类型", "描述", "效果", "备注"]
set_header(ws_events, events_headers)

events_data = [
    # 现有事件
    ("Core.Event.EncounterBandit", "遭遇战斗", "Combat", "路上遇到了强盗！", "gold:+20", "基础战斗事件"),
    ("Core.Event.JungleAmbush", "丛林伏击", "Combat", "丛林中突然冲出野兽！", "gold:+35,hp:-15", "有HP损失的战斗"),
    ("Core.Event.MountainBanditRaid", "山贼袭击", "Combat", "山贼拦路抢劫！", "gold:+50", "高收益战斗"),
    ("Core.Event.EncounterCaravan", "路遇商队", "Trade", "遇到了一支商队，可以进行交易。", "gold:-30,flag:has_traded", "交易事件"),
    ("Core.Event.MarketTrading", "集市贸易", "Trade", "来到了热闹的集市！", "gold:-50,item:HealthPotion:2", "购买道具"),
    ("Core.Event.AncientChest", "古旧宝箱", "Discovery", "路边发现了一个古旧的宝箱！", "gold:+100,item:AncientArtifact:1", "发现奖励"),
    ("Core.Event.HerbPatch", "草药丛", "Discovery", "发现了一片茂密的草药丛！", "item:HealthPotion:3", "野外采集"),
    ("Core.Event.UnexpectedEncounter", "随机遭遇", "Random", "旅途中发生了意想不到的事情...", "gold:+25", "随机收益"),
    ("Core.Event.KindStranger", "好心人", "Random", "遇到好心人送了一些盘缠！", "gold:+40", "随机友善事件"),
    ("Core.Event.AncientRuins", "神秘遗迹", "Story", "发现了一处被遗忘的神秘遗迹...", "flag:found_ruins,gold:+80", "剧情触发"),
    # 拓展事件
    ("Core.Event.TempleBlessing", "寺庙祈福", "Story", "路过一座古寺，入内参拜", "hp:+30,gold:-20", "拓展设计 - 付费恢复"),
    ("Core.Event.HiddenCave", "隐秘山洞", "Discovery", "发现了一个隐蔽的山洞入口", "item:GoldCoin:50,flag:found_cave", "拓展设计 - 隐藏地点"),
    ("Core.Event.Pickpocket", "遭遇扒手", "Combat", "可恶的扒手盯上了你！", "gold:-50,combat", "拓展设计 - 负面事件"),
    ("Core.Event.LostMerchant", "迷路商人", "Trade", "一位商人在路边摆摊，商品齐全", "gold:+100,item:random", "拓展设计 - 特殊商店"),
    ("Core.Event.WildAnimal", "驯服野兽", "Random", "遇到一只受伤的野生动物", "pet_chance:20%,gold:-30", "拓展设计 - 宠物机会"),
    ("Core.Event.Thunderstorm", "暴风雨", "Random", "突然来袭的暴风雨", "travel_speed:-50%,hp:-10", "拓展设计 - 天气影响"),
    ("Core.Event.Starfall", "流星雨", "Random", "罕见的流星雨划过夜空", "exp:+50,rare_item_chance", "拓展设计 - 稀有事件"),
]

for i, event in enumerate(events_data, 2):
    is_expansion = "拓展设计" in str(event)
    set_row(ws_events, event[:5], i, is_expansion=is_expansion)
    if len(event) > 5:
        add_expansion_marker(ws_events, i, 5, event[5])

auto_width(ws_events)

# ============================================================
# 工作表6: 事件树表 (EventTrees)
# ============================================================
ws_trees = wb.create_sheet("事件树表")

trees_headers = ["ID", "名称", "描述", "根节点", "节点数", "分支数", "结局数", "备注"]
set_header(ws_trees, trees_headers)

trees_data = [
    # 现有事件树
    ("Core.EventTree.MerchantEncounter", "路遇商队", "旅行途中遇到一支商队，可以选择交易或聊天", "node_1", 5, 3, 2, "商队分支叙事"),
    ("Core.EventTree.AncientRuins", "神秘遗迹", "发现了一处被遗忘的神秘遗迹，可以选择探索或离开", "node_1", 6, 3, 3, "遗迹探索分支叙事"),
    ("Core.EventTree.InnRest", "路旁旅店", "天色渐晚，发现了一家路边旅店，可以选择休息或继续赶路", "node_1", 4, 3, 2, "旅店休息分支叙事"),
    # 拓展事件树
    ("Core.EventTree.DragonLair", "巨龙巢穴", "在山洞深处发现了巨龙的踪迹...", "node_1", 8, 4, 3, "拓展设计 - BOSS战事件树"),
    ("Core.EventTree.TreasureHunt", "宝藏猎人", "根据古老的地图，你来到了一处疑似宝藏地点", "node_1", 10, 4, 4, "拓展设计 - 解谜探险事件树"),
    ("Core.EventTree.Tournament", "武道大会", "路过一座正在举办武道大会的城市", "node_1", 7, 3, 3, "拓展设计 - 竞技事件树"),
    ("Core.EventTree.MysteriousTrader", "神秘商人", "深夜，一位神秘的商人出现在你面前", "node_1", 6, 3, 3, "拓展设计 - 神秘交易事件树"),
]

for i, tree in enumerate(trees_data, 2):
    is_expansion = "拓展设计" in str(tree)
    set_row(ws_trees, tree[:7], i, is_expansion=is_expansion)
    if len(tree) > 7:
        add_expansion_marker(ws_trees, i, 7, tree[7])

auto_width(ws_trees)

# ============================================================
# 工作表7: 职业表 (Jobs)
# ============================================================
ws_jobs = wb.create_sheet("职业表")

jobs_headers = ["ID", "名称", "主属性", "战斗加成", "特殊效果", "解锁条件", "备注"]
set_header(ws_jobs, jobs_headers)

jobs_data = [
    # 现有职业
    ("Core.Job.Merchant", "商贾", "魅力", "交易价格+20%", "议价技能", "初始解锁", "交易型职业"),
    ("Core.Job.Escort", "镖师", "攻击", "伤害+15%", "护卫技能", "初始解锁", "战斗型职业"),
    ("Core.Job.Scholar", "学者", "智慧", "事件触发+30%", "鉴定技能", "初始解锁", "探索型职业"),
    ("Core.Job.Healer", "医者", "体力", "治疗效果+50%", "治疗技能", "初始解锁", "辅助型职业"),
    # 拓展职业
    ("Core.Job.Thief", "盗贼", "速度", "闪避+25%", "偷窃技能", "完成盗贼任务", "拓展设计 - 特殊职业"),
    ("Core.Job.Paladin", "圣武士", "防御", "减伤+20%", "神圣护盾", "完成神殿任务", "拓展设计 - 防护职业"),
    ("Core.Job.Ranger", "游侠", "速度", "远程+25%", "陷阱感知", "完成野外生存任务", "拓展设计 - 侦察职业"),
    ("Core.Job.Alchemist", "炼金师", "智慧", "药剂效果+40%", "自制药剂", "完成炼金任务", "拓展设计 - 生产职业"),
    ("Core.Job.Bard", "吟游诗人", "魅力", "队伍经验+15%", "鼓舞技能", "收集10首乐曲", "拓展设计 - 音乐职业"),
    ("Core.Job.Necromancer", "亡灵法师", "智慧", "召唤物+50%", "亡灵召唤", "收集10个灵魂", "拓展设计 - 黑暗职业"),
]

for i, job in enumerate(jobs_data, 2):
    is_expansion = "拓展设计" in str(job)
    set_row(ws_jobs, job[:6], i, is_expansion=is_expansion)
    if len(job) > 6:
        add_expansion_marker(ws_jobs, i, 6, job[6])

auto_width(ws_jobs)

# ============================================================
# 工作表8: 装备槽位表 (EquipmentSlots)
# ============================================================
ws_slots = wb.create_sheet("装备槽位表")

slots_headers = ["槽位ID", "名称", "装备类型", "稀有度影响", "特殊属性", "解锁条件", "备注"]
set_header(ws_slots, slots_headers)

slots_data = [
    # 现有及设计槽位
    ("Core.EquipSlot.Weapon", "武器", "单手/双手/远程", "伤害值", "攻击速度,暴击率", "初始解锁", "主要输出装备"),
    ("Core.EquipSlot.Armor", "护甲", "轻甲/中甲/重甲", "防御值+速度惩罚", "生命值,护甲穿透抗性", "初始解锁", "主要防御装备"),
    ("Core.EquipSlot.Accessory1", "饰品1", "戒指/项链/护符", "特殊属性", "各类属性加成", "初始解锁", "自由属性槽位"),
    ("Core.EquipSlot.Accessory2", "饰品2", "戒指/项链/护符", "特殊属性", "各类属性加成", "初始解锁", "自由属性槽位"),
    ("Core.EquipSlot.Mount", "坐骑", "马/骆驼/马车", "旅行速度+容量", "背包容量,脚程效率", "10级解锁", "旅行相关装备"),
    # 拓展槽位
    ("Core.EquipSlot.Helmet", "头盔", "轻/中/重", "防御+属性", "生命值,魔抗", "5级解锁", "拓展设计 - 新槽位"),
    ("Core.EquipSlot.Boots", "靴子", "轻/中/重", "速度+防御", "移动速度,闪避率", "8级解锁", "拓展设计 - 新槽位"),
    ("Core.EquipSlot.Gloves", "手套", "轻/中/重", "攻击+防御", "攻击速度,命中精度", "12级解锁", "拓展设计 - 新槽位"),
    ("Core.EquipSlot.Cloak", "披风", "布料/皮甲", "特殊抗性", "火抗,冰抗,雷抗", "15级解锁", "拓展设计 - 元素抗性"),
    ("Core.EquipSlot.Belt", "腰带", "皮革/金属", "生命+容量", "背包容量,负重上限", "10级解锁", "拓展设计 - 仓储扩展"),
]

for i, slot in enumerate(slots_data, 2):
    is_expansion = "拓展设计" in str(slot)
    set_row(ws_slots, slot[:6], i, is_expansion=is_expansion)
    if len(slot) > 6:
        add_expansion_marker(ws_slots, i, 6, slot[6])

auto_width(ws_slots)

# ============================================================
# 工作表9: 阵营表 (Affiliations)
# ============================================================
ws_aff = wb.create_sheet("阵营表")

aff_headers = ["阵营ID", "名称", "颜色", "可交互", "可战斗", "可交易", "可招募", "敌对阵营", "备注"]
set_header(ws_aff, aff_headers)

aff_data = [
    ("Core.Affiliation.Player", "玩家", "N/A", "是", "是", "是", "是", "Hostile,Authority", "玩家自身阵营"),
    ("Core.Affiliation.Friendly", "友善", "绿色(0.2,0.8,0.2)", "是", "否", "是", "是", "Hostile", "会帮助玩家的NPC"),
    ("Core.Affiliation.Allied", "友方", "蓝色(0.2,0.5,0.9)", "是", "否", "是", "是", "Hostile", "同盟关系的NPC"),
    ("Core.Affiliation.Neutral", "中立", "灰色(0.7,0.7,0.7)", "是", "可选", "是", "否", "-", "无特殊关系的NPC"),
    ("Core.Affiliation.Hostile", "敌对", "红色(0.9,0.2,0.2)", "否", "是", "否", "否", "Player,Friendly,Allied", "敌人和野兽"),
    ("Core.Affiliation.Authority", "权威", "紫色(0.7,0.3,0.9)", "是", "特殊", "是", "否", "Hostile", "守卫、官员等"),
    # 拓展阵营
    ("Core.Affiliation.Guild", "公会", "金色(0.9,0.8,0.2)", "是", "否", "是", "是", "Hostile", "拓展设计 - 工会势力"),
    ("Core.Affiliation.Rebel", "叛军", "红色暗(0.6,0.1,0.1)", "是", "可选", "是", "否", "Authority", "拓展设计 - 起义势力"),
    ("Core.Affiliation.Criminal", "罪犯", "黑色(0.2,0.2,0.2)", "是", "是", "是", "否", "Authority,Guild", "拓展设计 - 地下势力"),
]

for i, aff in enumerate(aff_data, 2):
    is_expansion = "拓展设计" in str(aff)
    set_row(ws_aff, aff[:8], i, is_expansion=is_expansion)
    if len(aff) > 8:
        add_expansion_marker(ws_aff, i, 8, aff[8])

auto_width(ws_aff)

# ============================================================
# 工作表10: 节点类型表 (LocationTypes)
# ============================================================
ws_locations = wb.create_sheet("节点类型表")

loc_headers = ["类型ID", "名称", "出现概率", "事件类型", "预期收益", "旅行时间", "备注"]
set_header(ws_locations, loc_headers)

loc_data = [
    ("Core.Location.Start", "起点", "固定", "无", "-", "-", "旅程起点"),
    ("Core.Location.City", "古城", "15%", "NPC遭遇/商店", "中等金币/道具", "中", "安全区域"),
    ("Core.Location.Market", "集市", "15%", "交易事件", "低但稳定", "短", "商业区域"),
    ("Core.Location.Wilderness", "荒野", "20%", "战斗/随机", "高波动", "中", "危险区域"),
    ("Core.Location.Dungeon", "副本", "25%", "连续战斗", "高但耗时", "长", "挑战区域"),
    ("Core.Location.Boss", "BOSS", "10%", "极限战斗", "极高一次性", "长", "BOSS区域"),
    ("Core.Location.Event", "事件点", "10%", "分支叙事", "极高看选择", "中", "剧情区域"),
    ("Core.Location.Goal", "终点", "固定", "阶段结算", "-", "-", "旅程终点"),
    # 拓展节点
    ("Core.Location.Shrine", "神殿", "5%", "祈福/任务", "buff/任务", "短", "拓展设计 - 宗教区域"),
    ("Core.Location.Cave", "山洞", "12%", "探索/宝藏", "高风险高回报", "中", "拓展设计 - 探险区域"),
    ("Core.Location.Village", "村庄", "10%", "休息/交易", "恢复/补给", "短", "拓展设计 - 友好区域"),
    ("Core.Location.Ruins", "遗迹", "8%", "考古/陷阱", "古代遗物", "长", "拓展设计 - 危险区域"),
    ("Core.Location.Camp", "营地", "15%", "遭遇/休息", "随机事件", "短", "拓展设计 - 中立区域"),
]

for i, loc in enumerate(loc_data, 2):
    is_expansion = "拓展设计" in str(loc)
    set_row(ws_locations, loc[:6], i, is_expansion=is_expansion)
    if len(loc) > 6:
        add_expansion_marker(ws_locations, i, 6, loc[6])

auto_width(ws_locations)

# ============================================================
# 工作表11: 物品类型表 (ItemTypes)
# ============================================================
ws_types = wb.create_sheet("物品类型表")

type_headers = ["类型ID", "名称", "可堆叠", "可丢弃", "可交易", "可使用", "关键属性", "备注"]
set_header(ws_types, type_headers)

type_data = [
    ("Core.ItemType.Money", "货币", "是", "否", "是", "否", "-", "游戏基础货币"),
    ("Core.ItemType.Food", "食物", "是", "是", "是", "是", "foodCalorific(热量)", "可食用恢复体力"),
    ("Core.ItemType.Weapon", "武器", "否", "是", "是", "否", "damage(伤害)", "装备增加攻击"),
    ("Core.ItemType.Armor", "护甲", "否", "是", "是", "否", "armor(防御)", "装备增加防御"),
    ("Core.ItemType.Consumable", "消耗品", "是", "是", "是", "是", "effect(效果)", "使用后消耗"),
    ("Core.ItemType.Material", "材料", "是", "是", "是", "否", "-", "用于制造和交易"),
    ("Core.ItemType.QuestItem", "任务物品", "否", "否", "否", "否", "-", "不可丢弃/交易"),
    # 拓展类型
    ("Core.ItemType.Accessory", "饰品", "否", "是", "是", "否", "特殊属性", "拓展设计 - 新类型"),
    ("Core.ItemType.Mount", "坐骑", "否", "是", "是", "否", "速度加成,容量加成", "拓展设计 - 旅行装备"),
    ("Core.ItemType.Book", "书籍", "是", "是", "是", "是", "经验/智慧加成", "拓展设计 - 知识道具"),
    ("Core.ItemType.Key", "钥匙", "否", "否", "否", "是", "开启特定门/宝箱", "拓展设计 - 解谜道具"),
    ("Core.ItemType.Seed", "种子", "是", "是", "是", "是", "种植产出", "拓展设计 - 种植系统"),
]

for i, item_type in enumerate(type_data, 2):
    is_expansion = "拓展设计" in str(item_type)
    set_row(ws_types, item_type[:7], i, is_expansion=is_expansion)
    if len(item_type) > 7:
        add_expansion_marker(ws_types, i, 7, item_type[7])

auto_width(ws_types)

# ============================================================
# 保存文件
# ============================================================
output_path = r"E:\UnityProgram\Game1\Docs\AGENT策划\list.xlsx"
wb.save(output_path)
print(f"Excel文件已保存到: {output_path}")
print(f"共创建 {len(wb.sheetnames)} 个工作表:")
for name in wb.sheetnames:
    print(f"  - {name}")
