# -*- coding: utf-8 -*-
"""
Game1 项目数据导出工具 - 中华文化扩展版
基于网络搜索结果添加中华风格内容
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 打开已存在的工作簿
wb = openpyxl.load_workbook(r"E:\UnityProgram\Game1\Docs\AGENT策划\list.xlsx")

# 定义样式
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
expansion_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
epic_fill = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")  # 淡紫色

def set_header(ws, headers, row=1):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

def set_row(ws, data, row, col_start=1, is_expansion=False, is_epic=False):
    for i, value in enumerate(data, col_start):
        cell = ws.cell(row=row, column=i, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border
        if is_epic:
            cell.fill = epic_fill
        elif is_expansion:
            cell.fill = expansion_fill

def add_expansion_marker(ws, row, total_cols, note=""):
    cell = ws.cell(row=row, column=total_cols + 1, value=f"[拓展] {note}")
    cell.fill = expansion_fill
    cell.alignment = center_alignment

def auto_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

# ============================================================
# 新增工作表: 中华药材表 (ChineseHerbs)
# ============================================================
ws_herbs = wb.create_sheet("中华药材")

herbs_headers = ["ID", "名称", "类别", "效果", "稀有度", "售价", "备注"]
set_header(ws_herbs, herbs_headers)

herbs_data = [
    # 普通药材
    ("Core.Herb.Ginseng", "人参", "上品药材", "恢复30HP,短暂内攻击+5%", "稀有", 500, "补气养血"),
    ("Core.Herb.Lingzhi", "灵芝", "上品药材", "恢复20HP,防御+10%持续3回合", "稀有", 800, "仙草"),
    ("Core.Herb.SnowLotus", "天山雪莲", "极品药材", "解除所有debuff,HP全满", "史诗", 2000, "雪山珍品"),
    ("Core.Herb.Cordyceps", "冬虫夏草", "上品药材", "恢复40HP,速度+15%持续2回合", "稀有", 1200, "滋补圣品"),
    ("Core.Herb.Honey", "蜂蜜", "普通药材", "恢复15HP,攻击力+5%持续1回合", "普通", 50, "天然甜味"),
    ("Core.Herb.GojiBerry", "枸杞", "普通药材", "恢复10HP,暴击率+3%持续2回合", "普通", 30, "养生佳品"),
    ("Core.Herb.Rehmania", "熟地黄", "普通药材", "恢复25HP,持续恢复5HP/秒持续10秒", "普通", 80, "补血药材"),
    ("Core.Herb.Saffron", "藏红花", "上品药材", "恢复15HP,全体治疗10HP", "稀有", 1500, "活血化瘀"),
    ("Core.Herb.Licorice", "甘草", "普通药材", "恢复10HP,所有治疗效果+10%", "普通", 20, "百搭药材"),
    # 丹药类
    ("Core.Pill.SmallRecovery", "小还丹", "下品丹药", "恢复50HP", "普通", 200, "入门丹药"),
    ("Core.Pill.LargeRecovery", "大还丹", "中品丹药", "恢复150HP", "稀有", 800, "进阶丹药"),
    ("Core.Pill.Xiaoyao", "逍遥丹", "上品丹药", "恢复100HP,解除中毒", "稀有", 1200, "解毒神药"),
    ("Core.Pill.ReturnSoul", "返魂香", "极品丹药", "死亡时复活,恢复50%HP", "史诗", 5000, "救命仙丹"),
    ("Core.Pill.NineTurn", "九转还魂丹", "神品丹药", "HP全满,所有属性+20%持续5回合", "传说", 10000, "至高神药"),
    # 拓展药材
    ("Core.Herb.HeShouWu", "何首乌", "上品药材", "永久最大HP+5", "稀有", 600, "拓展设计 - 乌发益精"),
    ("Core.Herb.WuWeiZi", "五味子", "普通药材", "恢复20HP,内力+30", "普通", 100, "拓展设计 - 五味俱全"),
    ("Core.Herb.DongZhi", "天冬", "普通药材", "恢复15HP,闪避率+5%持续2回合", "普通", 60, "拓展设计 - 滋阴润燥"),
    ("Core.Herb.ChuanXiong", "川芎", "普通药材", "恢复10HP,攻击速度+10%", "普通", 40, "拓展设计 - 活血行气"),
]

row = 2
for herb in herbs_data:
    is_expansion = "拓展设计" in str(herb)
    is_epic = herb[4] == "史诗" or herb[4] == "传说"
    set_row(ws_herbs, herb[:6], row, is_expansion=is_expansion, is_epic=is_epic)
    if len(herb) > 6:
        add_expansion_marker(ws_herbs, row, 6, herb[6])
    row += 1

auto_width(ws_herbs)

# ============================================================
# 新增工作表: 中华矿石表 (ChineseOres)
# ============================================================
ws_ores = wb.create_sheet("中华矿石")

ores_headers = ["ID", "名称", "等级", "用途", "售价", "产地", "备注"]
set_header(ws_ores, ores_headers)

ores_data = [
    # 普通矿石
    ("Core.Ore.IronOre", "铁矿石", 1, "锻造基础装备", 10, "铁矿", "最常见矿石"),
    ("Core.Ore.CopperOre", "铜矿石", 1, "锻造青铜装备", 15, "铜矿", "初级金属"),
    ("Core.Ore.SilverOre", "银矿石", 2, "锻造银器/货币", 50, "银矿", "贵金属"),
    ("Core.Ore.GoldOre", "金矿石", 3, "锻造金器/货币", 200, "金矿", "贵金属"),
    # 玉石类
    ("Core.Ore.Soapstone", "羊脂玉", 4, "顶级饰品材料", 500, "玉矿", "温润如脂"),
    ("Core.Ore.HetianJade", "和田玉", 5, "传说饰品材料", 2000, "新疆和田", "玉中上品"),
    ("Core.Ore.LapisLazuli", "蓝纹玉", 4, "高级饰品材料", 800, "昆仑山", "蓝色纹理"),
    ("Core.Ore.TianheStone", "天河石", 3, "中级饰品材料", 300, "南方矿区", "微透明"),
    ("Core.Ore.ShoushanStone", "寿山石", 4, "印章/雕件材料", 600, "福建寿山", "彩石"),
    # 宝石类
    ("Core.Gem.CatsEye", "猫眼石", 4, "镶嵌用宝石", 1000, "云南", "稀有宝石"),
    ("Core.Gem.Diamond", "金刚石", 5, "顶级镶嵌宝石", 3000, "西域", "硬度最高"),
    ("Core.Gem.Lapis", "青金石", 4, "蓝色宝石", 800, "阿富汗", "古代珍品"),
    ("Core.Gem.Obsidian", "黑曜石", 3, "暗属性宝石", 400, "火山地区", "辟邪之石"),
    ("Core.Gem.DushanStone", "独山玉", 3, "南阳特产", 350, "河南独山", "多色玉石"),
    # 灵石类
    ("Core.Spirit.StarlightJade", "星光石", 5, "注灵材料", 5000, "星空之地", "蕴含星力"),
    ("Core.Spirit.DragonSlumber", "龙涎石", 6, "传说注灵材料", 15000, "深海龙宫", "龙族遗珍"),
    ("Core.Spirit.NuwaStone", "女娲石", 7, "神话修复材料", 50000, "女娲遗迹", "补天之石"),
    # 拓展矿石
    ("Core.Ore.TinOre", "锡矿石", 1, "焊接材料", 12, "锡矿", "拓展设计"),
    ("Core.Ore.LeadOre", "铅矿石", 1, "工业材料", 8, "铅矿", "拓展设计"),
    ("Core.Ore.Jadeite", "翡翠", 5, "高级饰品", 2500, "缅甸", "拓展设计 - 翠绿欲滴"),
    ("Core.Ore.Ruby", "红宝石", 5, "镶嵌宝石", 2800, "云南", "拓展设计 - 鸽血红"),
    ("Core.Ore.Emerald", "祖母绿", 5, "镶嵌宝石", 3000, "哥伦比亚", "拓展设计 - 翠绿"),
    ("Core.Ore.Sapphire", "蓝宝石", 5, "镶嵌宝石", 2500, "斯里兰卡", "拓展设计 - 皇家蓝"),
]

row = 2
for ore in ores_data:
    is_expansion = "拓展设计" in str(ore)
    set_row(ws_ores, ore[:6], row, is_expansion=is_expansion)
    if len(ore) > 6:
        add_expansion_marker(ws_ores, row, 6, ore[6])
    row += 1

auto_width(ws_ores)

# ============================================================
# 新增工作表: 中华状态效果表 (ChineseStatus)
# ============================================================
ws_status = wb.create_sheet("中华状态效果")

status_headers = ["ID", "名称", "类型", "持续时间", "效果", "解除方法", "备注"]
set_header(ws_status, status_headers)

status_data = [
    # 中毒类
    ("Core.Status.HotPoison", "热毒", "Debuff", "3回合", "每回合-8HP", "清热药材", "热邪入体"),
    ("Core.Status.ColdPoison", "寒毒", "Debuff", "3回合", "每回合-5HP,速度-20%", "温阳药材", "寒邪侵体"),
    ("Core.Status.RotPoison", "腐毒", "Debuff", "4回合", "每回合-10HP", "解毒丹", "腐蚀血肉"),
    ("Core.Status.GuPoison", "蛊毒", "Debuff", "5回合", "每回合-3HP,吸收营养", "牛黄丸", "蛊虫侵蚀"),
    # 增益类
    ("Core.Status.QiFocus", "聚气", "Buff", "3回合", "攻击+25%", "自然消散", "真气凝聚"),
    ("Core.Status.BloodRush", "血气方刚", "Buff", "2回合", "暴击率+15%", "自然消散", "气血充盈"),
    ("Core.Status.GoldenBell", "金钟罩", "Buff", "2回合", "受到伤害-30%", "受到攻击", "硬气功"),
    ("Core.Status.IronShirt", "铁布衫", "Buff", "3回合", "防御+20%", "自然消散", "横练功夫"),
    ("Core.Status.SwiftWind", "风步", "Buff", "2回合", "速度+40%,闪避+10%", "受到攻击", "轻功"),
    ("Core.Status.EagleEye", "鹰眼", "Buff", "3回合", "命中率+20%,暴击伤害+30%", "自然消散", "目力过人"),
    # 减益类
    ("Core.Status.Dizziness", "眩晕", "Debuff", "1回合", "无法行动", "受到治疗", "头目晕眩"),
    ("Core.Status.Freeze", "冰封", "Debuff", "1回合", "无法行动,受伤+20%", "火系攻击", "寒冰入体"),
    ("Core.Status.Paralysis", "麻痹", "Debuff", "2回合", "速度-50%,攻击停顿", "活血药材", "经脉不通"),
    ("Core.Status.Blind", "致盲", "Debuff", "2回合", "命中率-50%", "明目丹", "双目失明"),
    ("Core.Status.Silence", "沉默", "Debuff", "2回合", "无法使用技能", "清心丹", "禁锢内力"),
    ("Core.Status.Terror", "恐惧", "Debuff", "2回合", "有30%概率逃跑", "安神丸", "心惊胆战"),
    # 特殊类
    ("Core.Status.Burning", "灼烧", "Debuff", "2回合", "每回合-15HP", "冷水", "火焰持续"),
    ("Core.Status.Bleeding", "流血", "Debuff", "3回合", "每回合-5HP,可叠加5层", "止血草", "伤口撕裂"),
    ("Core.Status.Weakness", "虚弱", "Debuff", "3回合", "攻击力-30%", "补气汤", "体力不支"),
    ("Core.Status.Slow", "迟缓", "Debuff", "2回合", "速度-30%", "麻黄", "动作迟缓"),
    # 拓展状态
    ("Core.Status.Curse", "诅咒", "Debuff", "永久", "全属性-10%", "净化符", "拓展设计 - 邪术"),
    ("Core.Status.Blessing", "祝福", "Buff", "5回合", "全属性+15%", "自然消散", "拓展设计 - 神佑"),
    ("Core.Status.Invincible", "金刚不坏", "Buff", "1回合", "免疫所有伤害", "受到攻击", "拓展设计 - 至高防御"),
]

row = 2
for status in status_data:
    is_expansion = "拓展设计" in str(status)
    set_row(ws_status, status[:6], row, is_expansion=is_expansion)
    if len(status) > 6:
        add_expansion_marker(ws_status, row, 6, status[6])
    row += 1

auto_width(ws_status)

# ============================================================
# 新增工作表: 中华怪物表 (ChineseMonsters)
# ============================================================
ws_monsters = wb.create_sheet("中华怪物")

monsters_headers = ["ID", "名称", "类型", "等级", "HP", "攻击", "防御", "速度", "掉落", "备注"]
set_header(ws_monsters, monsters_headers)

monsters_data = [
    # 妖兽类
    ("Core.Monster.FoxSpirit", "狐妖", "妖兽", 10, 80, 15, 10, 1.2, "妖核×2,金币×50", "九尾妖狐眷族"),
    ("Core.Monster.WhiteFox", "白狐", "妖兽", 8, 60, 12, 8, 1.3, "狐毛×3,金币×30", "灵性极高"),
    ("Core.Monster.SnakeSpirit", "蛇妖", "妖兽", 12, 100, 18, 12, 1.0, "蛇胆×2,金币×60", "修炼成精"),
    ("Core.Monster.Ghost", "孤魂野鬼", "鬼魂", 5, 40, 20, 2, 1.5, "幽魂×1,金币×20", "怨念不散"),
    ("Core.Monster.FierceGhost", "厉鬼", "鬼魂", 15, 120, 25, 5, 1.4, "厉魂×2,金币×100", "含冤而死"),
    ("Core.Monster.Zombie", "跳尸", "僵尸", 18, 200, 22, 25, 0.5, "僵尸心×1,金币×80", "白僵进化"),
    ("Core.Monster.TreeSpirit", "树妖", "精怪", 20, 150, 18, 30, 0.8, "千年古木×1,金币×120", "草木成精"),
    ("Core.Monster.JadeBear", "玉麒麟", "神兽", 30, 500, 40, 35, 1.0, "麒麟角×1,金币×500", "祥瑞之兽"),
    # 山贼类
    ("Core.Monster.Bandit", "山贼", "匪徒", 8, 50, 10, 8, 0.9, "朴刀×1,金币×25", "占山为王"),
    ("Core.Monster.RogueLeader", "草寇头目", "匪徒", 12, 100, 15, 12, 0.8, "皮甲×1,金币×60", "小股势力"),
    ("Core.Monster流动寇", "流寇", "匪徒", 10, 70, 14, 10, 1.1, "金币×40,随机物品", "流窜作案"),
    ("Core.Monster.LocalBully", "恶霸", "匪徒", 15, 120, 20, 18, 0.7, "保护费×100,金币×80", "地方一霸"),
    # 神话生物
    ("Core.Monster.Asura", "修罗", "神话", 25, 300, 45, 20, 1.3, "修罗之血×1,金币×300", "阿修罗族"),
    ("Core.Monster.Yaksha", "夜叉", "神话", 22, 250, 40, 25, 1.2, "夜叉皮×1,金币×250", "鬼怪生物"),
    ("Core.Monster.Rakshasa", "罗刹", "神话", 28, 350, 50, 22, 1.4, "罗刹之牙×1,金币×350", "恶鬼"),
    # 拓展怪物
    ("Core.Monster.Scorpion", "毒蝎", "妖兽", 6, 45, 18, 6, 0.7, "蝎尾×2,金币×25", "拓展设计 - 沙漠"),
    ("Core.Monster.Mantis", "螳螂精", "妖兽", 9, 65, 22, 8, 1.6, "螳臂×1,金币×40", "拓展设计 - 速度快"),
    ("Core.Monster.Spider", "蛛妖", "妖兽", 14, 130, 16, 20, 0.9, "蛛丝×3,金币×70", "拓展设计 - 吐丝束缚"),
    ("Core.Monster.WolfKing", "狼王", "妖兽", 20, 180, 30, 15, 1.5, "狼牙×2,金币×150", "拓展设计 - 群狼首领"),
    ("Core.Monster.BlackTortoise", "玄龟", "神兽", 25, 400, 20, 50, 0.3, "龟甲×2,金币×200", "拓展设计 - 极高防御"),
    ("Core.Monster.Qilin", "麒麟", "神兽", 35, 600, 45, 40, 1.1, "麒麟鳞×2,金币×800", "拓展设计 - 祥瑞"),
]

row = 2
for monster in monsters_data:
    is_expansion = "拓展设计" in str(monster)
    set_row(ws_monsters, monster[:9], row, is_expansion=is_expansion)
    if len(monster) > 9:
        add_expansion_marker(ws_monsters, row, 9, monster[9])
    row += 1

auto_width(ws_monsters)

# ============================================================
# 新增工作表: 中华武器表 (ChineseWeapons)
# ============================================================
ws_weapons = wb.create_sheet("中华武器")

weapons_headers = ["ID", "名称", "类型", "等级", "伤害", "特效", "职业", "备注"]
set_header(ws_weapons, weapons_headers)

weapons_data = [
    # 刀类
    ("Core.Weapon.SimpleBlade", "朴刀", "刀", 1, 8, "-", "通用", "基础武器"),
    ("Core.Weapon.BroadBlade", "戒刀", "刀", 5, 15, "暴击+5%", "通用", "佛门武器"),
    ("Core.Weapon.MoonBlade", "偃月刀", "刀", 15, 35, "范围伤害+10%", "镖师", "长柄大刀"),
    ("Core.Weapon.DragonSaber", "龙牙刀", "刀", 25, 55, "对龙系伤害+30%", "刺客", "传说武器"),
    # 剑类
    ("Core.Weapon.ShortSword", "短剑", "剑", 1, 6, "速度+10%", "刺客", "入门剑器"),
    ("Core.Weapon.LongSword", "长剑", "剑", 5, 12, "-", "通用", "标准武器"),
    ("Core.Weapon.FineSword", "精钢剑", "剑", 10, 22, "破甲+5", "镖师", "精制武器"),
    ("Core.Weapon.SpiritSword", "灵剑", "剑", 20, 45, "内力加成+15%", "学者", "蕴含灵气"),
    ("Core.Weapon.FrostBlade", "寒冰剑", "剑", 22, 48, "冰冻几率15%", "刺客", "冰系附魔"),
    # 枪类
    ("Core.Weapon.Spear", "长枪", "枪", 3, 10, "攻击距离+1", "镖师", "基础枪兵"),
    ("Core.Weapon.Halberd", "戟", "枪", 12, 28, "刺穿+10%", "镖师", "战阵武器"),
    ("Core.Weapon.DragonHalberd", "龙戟", "枪", 25, 58, "对BOSS伤害+25%", "将军", "战场之王"),
    # 扇类
    ("Core.Weapon.FoldingFan", "折扇", "扇", 2, 3, "敏捷+5%", "商贾", "文人武器"),
    ("Core.Weapon.WarFan", "羽扇", "扇", 15, 20, "技能冷却-10%", "智者", "诸葛亮之扇"),
    ("Core.Weapon.ImmortalFan", "仙羽扇", "扇", 28, 35, "全体攻击+20%", "医者", "仙家之物"),
    # 拓展武器
    ("Core.Weapon.Pudao", "扑刀", "刀", 8, 18, "连击+1", "镖师", "拓展设计 - 快刀"),
    ("Core.Weapon.WhipSword", "鞭剑", "剑", 18, 38, "灵活连击", "刺客", "拓展设计 - 软剑"),
    ("Core.Weapon.DoubleSword", "双剑", "剑", 20, 42, "双倍攻击次数", "刺客", "拓展设计 - 双武"),
    ("Core.Weapon.ThreeSection", "三节棍", "棍", 12, 25, "击晕几率10%", "镖师", "拓展设计 - 软兵"),
    ("Core.Weapon.Staff", "法杖", "杖", 10, 8, "技能伤害+25%", "医者", "拓展设计 - 法师武器"),
]

row = 2
for weapon in weapons_data:
    is_expansion = "拓展设计" in str(weapon)
    set_row(ws_weapons, weapon[:7], row, is_expansion=is_expansion)
    if len(weapon) > 7:
        add_expansion_marker(ws_weapons, row, 7, weapon[7])
    row += 1

auto_width(ws_weapons)

# ============================================================
# 新增工作表: 中华防具表 (ChineseArmor)
# ============================================================
ws_armor = wb.create_sheet("中华防具")

armor_headers = ["ID", "名称", "类型", "等级", "防御", "特效", "速度惩罚", "备注"]
set_header(ws_armor, armor_headers)

armor_data = [
    # 布甲类
    ("Core.Armor.ClothRobe", "布衣", "布甲", 1, 2, "魔抗+3", "0%", "基础防具"),
    ("Core.Armor.SilkRobe", "绸袍", "布甲", 5, 5, "魔抗+8,魅力+2", "0%", "商人常服"),
    ("Core.Armor.TaoistRobe", "道袍", "布甲", 12, 10, "内力回复+20%", "5%", "道士装备"),
    # 皮甲类
    ("Core.Armor.LeatherArmor", "皮甲", "轻甲", 3, 8, "-", "10%", "基础轻甲"),
    ("Core.Armor.HardLeather", "硬皮甲", "轻甲", 8, 15, "防御+5", "15%", "正规军装"),
    ("Core.Armor.JadeLeather", "玉皮甲", "轻甲", 18, 30, "所有抗性+10", "20%", "精良皮甲"),
    # 锁甲类
    ("Core.Armor.ChainMail", "锁子甲", "中甲", 6, 12, "-", "20%", "基础中甲"),
    ("Core.Armor.IronChainMail", "铁锁甲", "中甲", 14, 25, "格挡+10%", "25%", "正规军备"),
    ("Core.Armor.SilverChainMail", "银锁甲", "中甲", 22, 40, "圣光抗性+20", "30%", "骑士标配"),
    # 板甲类
    ("Core.Armor.PlateArmor", "板甲", "重甲", 15, 35, "免疫穿刺", "40%", "重装战士"),
    ("Core.Armor.DragonScale", "龙鳞甲", "重甲", 28, 60, "火抗+50,HP+100", "45%", "龙鳞打造"),
    ("Core.Armor.HeavenlyArmor", "天将甲", "重甲", 35, 80, "全属性+10%,无敌1秒/次", "50%", "神话级防具"),
    # 拓展防具
    ("Core.Armor.BambooArmor", "竹甲", "轻甲", 4, 6, "轻便+5%速度", "5%", "拓展设计 - 丛林"),
    ("Core.Armor.ScaledArmor", "鳞甲", "中甲", 10, 18, "穿刺抗性+15", "22%", "拓展设计 - 爬行类"),
    ("Core.Armor.MirrorArmor", "明光铠", "重甲", 25, 55, "反弹伤害15%", "42%", "拓展设计 - 镜面反射"),
    ("Core.Armor.GhostShroud", "幽魂披风", "布甲", 20, 15, "闪避+20%,暗属性", "8%", "拓展设计 - 鬼魅"),
]

row = 2
for armor in armor_data:
    is_expansion = "拓展设计" in str(armor)
    set_row(ws_armor, armor[:7], row, is_expansion=is_expansion)
    if len(armor) > 7:
        add_expansion_marker(ws_armor, row, 7, armor[7])
    row += 1

auto_width(ws_armor)

# ============================================================
# 新增工作表: 中华事件树详细表 (ChineseEventTrees)
# ============================================================
ws_event_tree = wb.create_sheet("中华事件树详细")

event_headers = ["事件ID", "节点ID", "节点类型", "标题", "描述", "选项数", "结局数", "备注"]
set_header(ws_event_tree, event_headers)

event_tree_data = [
    # 现有事件树扩展描述
    ("Core.EventTree.MerchantEncounter", "node_1~node_5", "Root/Choice/End", "路遇商队", "旅行途中遇到商队,选择交易或离开", 3, 2, "经典分支叙事"),
    ("Core.EventTree.AncientRuins", "node_1~node_6", "Root/Choice/End", "神秘遗迹", "探索古代遗迹,可能发现宝藏或陷阱", 3, 3, "探险分支叙事"),
    ("Core.EventTree.InnRest", "node_1~node_4", "Root/Choice/End", "路旁旅店", "旅途疲惫,在旅店休息或继续前进", 3, 2, "休息分支叙事"),
    # 拓展中华风事件树
    ("Core.EventTree.MysteriousMonk", "node_1~node_6", "Root/Choice/End", "神秘僧人", "拓展设计 - 路遇神秘僧人,可获得佛法加持或修行任务", 4, 3, "佛门事件"),
    ("Core.EventTree.DragonPool", "node_1~node_8", "Root/Random/Choice/End", "龙潭探宝", "拓展设计 - 传说龙潭,可能有真龙宝藏或龙族惩罚", 4, 4, "神话事件"),
    ("Core.EventTree.TeaHouse", "node_1~node_5", "Root/Choice/End", "茶馆奇遇", "拓展设计 - 古镇茶馆,听书了解秘密或遇到江湖人士", 3, 3, "江湖事件"),
    ("Core.EventTree.MedicineValley", "node_1~node_7", "Root/Choice/End", "药王谷", "拓展设计 - 误入药王谷,可采药或需解毒", 3, 4, "医药事件"),
    ("Core.EventTree.OverthrowBandits", "node_1~node_6", "Root/Choice/End", "剿灭山贼", "拓展设计 - 官兵请求协助剿匪,战斗与外交选择", 4, 3, "战斗事件"),
    ("Core.EventTree.GhostCity", "node_1~node_9", "Root/Choice/Random/End", "鬼城迷踪", "拓展设计 - 进入无人鬼城,探索真相或逃离", 5, 4, "鬼魂事件"),
    ("Core.EventTree.ImperialExam", "node_1~node_7", "Root/Choice/End", "赶考书生", "拓展设计 - 遇见赶考书生,可结伴或帮助", 3, 3, "书生事件"),
]

row = 2
for event in event_tree_data:
    is_expansion = "拓展设计" in str(event)
    set_row(ws_event_tree, event[:7], row, is_expansion=is_expansion)
    if len(event) > 7:
        add_expansion_marker(ws_event_tree, row, 7, event[7])
    row += 1

auto_width(ws_event_tree)

# ============================================================
# 新增工作表: 商贸商品详细表 (TradeGoodsDetail)
# ============================================================
ws_trade = wb.create_sheet("商贸商品详细")

trade_headers = ["商品ID", "名称", "产地", "收购价", "出售价", "利润%", "需求季节", "备注"]
set_header(ws_trade, trade_headers)

trade_data = [
    # 基础商品
    ("Core.Trade.Salt", "食盐", "盐井", 10, 18, "80%", "全年", "生活必需品"),
    ("Core.Trade.Rice", "大米", "稻田", 8, 14, "75%", "全年", "主食"),
    ("Core.Trade.Wheat", "小麦", "麦田", 6, 10, "67%", "秋季", "制作面粉"),
    # 丝绸之路商品
    ("Core.Trade.Silk", "丝绸", "江南", 100, 220, "120%", "春季", "丝绸之路核心"),
    ("Core.Trade.Tea", "茶叶", "茶山", 40, 90, "125%", "春季", "茶马古道"),
    ("Core.Trade.Porcelain", "瓷器", "景德镇", 150, 400, "167%", "全年", "外销主力"),
    # 香料类
    ("Core.Trade.Incense", "香料", "西域", 60, 150, "150%", "全年", "香料之路"),
    ("Core.Trade.Cinnamon", "肉桂", "南方", 30, 70, "133%", "全年", "调味料"),
    ("Core.Trade.StarAnise", "八角", "南方", 25, 55, "120%", "秋季", "炖肉调料"),
    # 药材类
    ("Core.Trade.Medicine", "药材", "各大药铺", 50, 120, "140%", "全年", "医商必备"),
    ("Core.Trade.Ginseng", "人参", "长白山", 300, 800, "167%", "秋季", "高档药材"),
    # 拓展商品
    ("Core.Trade.Antique", "古董", "遗迹", 200, 600, "200%", "不定", "拓展设计 - 鉴定价值"),
    ("Core.Trade.Jade", "玉石", "新疆", 400, 1200, "200%", "全年", "拓展设计 - 奢侈品"),
    ("Core.Trade.Horse", "马匹", "草原", 250, 500, "100%", "春季", "拓展设计 - 坐骑"),
    ("Core.Trade.Paper", "纸张", "造纸坊", 20, 50, "150%", "全年", "拓展设计 - 文人需求"),
    ("Core.Trade.Bamboo", "竹简", "竹林", 15, 40, "167%", "全年", "拓展设计 - 记录工具"),
    ("Core.Trade.Vine", "藤蔓", "南方", 10, 25, "150%", "夏季", "拓展设计 - 编织材料"),
]

row = 2
for trade in trade_data:
    is_expansion = "拓展设计" in str(trade)
    set_row(ws_trade, trade[:7], row, is_expansion=is_expansion)
    if len(trade) > 7:
        add_expansion_marker(ws_trade, row, 7, trade[7])
    row += 1

auto_width(ws_trade)

# ============================================================
# 新增工作表: 中华地名表 (ChinesePlaceNames)
# ============================================================
ws_places = wb.create_sheet("中华地名")

places_headers = ["地点ID", "名称", "类型", "区域", "特产", "事件", "备注"]
set_header(ws_places, places_headers)

places_data = [
    # 城市
    ("Core.Place.ChangAn", "长安城", "都城", "关中", "丝绸,瓷器,古籍", "科举,庙会,灯会", "大唐都城"),
    ("Core.Place.LuoYang", "洛阳城", "都城", "中原", "牡丹,古籍,药材", "花会,佛会", "千年帝都"),
    ("Core.Place.HangZhou", "杭州城", "商城", "江南", "茶叶,丝绸,龙井", "船宴,观潮", "人间天堂"),
    ("Core.Place.SuZhou", "苏州城", "商城", "江南", "丝绸,刺绣,园林", "评弹,昆曲", "江南水乡"),
    ("Core.Place.YangZhou", "扬州城", "商城", "江南", "盐业,玉石,美食", "盐商,瘦马", "淮左名都"),
    # 市镇
    ("Core.Place.JingZhou", "荆州城", "重镇", "湖北", "铁器,木材", "三国遗迹", "兵家必争"),
    ("Core.Place.XianYang", "咸阳城", "古镇", "关中", "农产品,皮草", "秦腔,古迹", "秦朝故都"),
    ("Core.Place.KaiFeng", "开封府", "都城", "中原", "美食,汴绣", "小吃,夜市", "北宋汴京"),
    # 特殊地点
    ("Core.Place.JadePass", "玉门关", "关卡", "西域", "玉石,香料", "丝路商队", "西域门户"),
    ("Core.Place.Yulin", "榆林", "边塞", "塞北", "战马,皮草,人参", "游牧交易", "塞外江南"),
    ("Core.Place.DuYan", "敦煌", "要塞", "西域", "壁画,佛像,绢画", "石窟艺术", "丝路明珠"),
    # 拓展地点
    ("Core.Place.MountWu", "武夷山", "山脉", "江南", "茶叶,草药", "采茶体验", "拓展设计 - 名山"),
    ("Core.Place.TaiShan", "泰山", "山脉", "山东", "石碑,药材", "封禅大典", "拓展设计 - 五岳之首"),
    ("Core.Place.WestLake", "西湖", "湖泊", "江南", "莲藕,鲤鱼,龙井", "画舫游湖", "拓展设计 - 人间仙境"),
    ("Core.Place.QinLing", "秦岭", "山脉", "关中", "珍禽,异兽,矿石", "深山探险", "拓展设计 - 天然宝库"),
    ("Core.Place.LiaoDong", "辽东", "边疆", "东北", "人参,貂皮,矿石", "游牧交易", "拓展设计 - 塞外"),
]

row = 2
for place in places_data:
    is_expansion = "拓展设计" in str(place)
    set_row(ws_places, place[:6], row, is_expansion=is_expansion)
    if len(place) > 6:
        add_expansion_marker(ws_places, row, 6, place[6])
    row += 1

auto_width(ws_places)

# ============================================================
# 保存文件
# ============================================================
output_path = r"E:\UnityProgram\Game1\Docs\AGENT策划\list.xlsx"
wb.save(output_path)
print(f"Excel文件已更新: {output_path}")
print(f"共 {len(wb.sheetnames)} 个工作表:")
for name in wb.sheetnames:
    print(f"  - {name}")
