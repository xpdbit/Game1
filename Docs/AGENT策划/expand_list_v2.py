# -*- coding: utf-8 -*-
"""
Game1 项目数据导出工具 - 进一步扩展版
添加更多拓展内容和中文名称优化
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 打开已存在的工作簿
wb = openpyxl.load_workbook(r"E:\UnityProgram\Game1\Docs\AGENT策划\list.xlsx")

# 定义样式
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
expansion_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
legendary_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

def set_header(ws, headers, row=1):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

def set_row(ws, data, row, col_start=1, is_expansion=False):
    for i, value in enumerate(data, col_start):
        cell = ws.cell(row=row, column=i, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border
        if is_expansion:
            cell.fill = expansion_fill

def add_expansion_marker(ws, row, total_cols, note=""):
    cell = ws.cell(row=row, column=total_cols + 1, value=f"[拓展] {note}")
    cell.fill = expansion_fill
    cell.alignment = center_alignment

def auto_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

# ============================================================
# 新增工作表: 商贸商品表 (TradeGoods)
# ============================================================
ws_goods = wb.create_sheet("商贸商品")

goods_headers = ["ID", "名称", "类别", "收购价", "出售价", "利润", "产地", "备注"]
set_header(ws_goods, goods_headers)

goods_data = [
    # 基础商品
    ("Core.Good.Salt", "食盐", "必需品", 10, 15, "50%", "盐井", "基础生活物资"),
    ("Core.Good.Rice", "大米", "必需品", 8, 12, "50%", "稻田", "基础粮食"),
    ("Core.Good.Tea", "茶叶", "奢侈品", 30, 50, "67%", "茶山", "高利润商品"),
    ("Core.Good.Silk", "丝绸", "奢侈品", 80, 150, "88%", "蚕桑", "顶级商品"),
    ("Core.Good.Porcelain", "瓷器", "艺术品", 100, 200, "100%", "瓷窑", "高价商品"),
    ("Core.Good.Tobacco", "烟草", "嗜好品", 25, 40, "60%", "烟田", "稳定需求"),
    ("Core.Good.Wine", "酒类", "嗜好品", 20, 35, "75%", "酒坊", "旅行消耗品"),
    # 中级商品
    ("Core.Good.Medicine", "药材", "必需品", 40, 70, "75%", "药铺", "医商必备"),
    ("Core.Good.Thread", "棉纱", "原料", 15, 25, "67%", "纺织", "制作原料"),
    ("Core.Good.Iron", "生铁", "原料", 25, 40, "60%", "铁矿", "锻造原料"),
    ("Core.Good.Wood", "木材", "原料", 12, 20, "67%", "林场", "建筑原料"),
    # 高级商品
    ("Core.Good.Jade", "玉石", "奢侈品", 200, 500, "150%", "玉矿", "顶级奢侈品"),
    ("Core.Good.Pearl", "珍珠", "奢侈品", 150, 400, "167%", "珍珠湾", "女性最爱"),
    ("Core.Good.Gold", "黄金", "货币", 500, 700, "40%", "金矿", "硬通货"),
    ("Core.Good.Silver", "白银", "货币", 100, 140, "40%", "银矿", "稳定货币"),
    # 拓展商品
    ("Core.Good.Incense", "香料", "奢侈品", 60, 120, "100%", "西域", "拓展设计 - 香料之路"),
    ("Core.Good.Antique", "古董", "艺术品", 300, 800, "167%", "遗迹", "拓展设计 - 鉴定价值"),
    ("Core.Good.Horse", "马匹", "交通工具", 200, 350, "75%", "草原", "拓展设计 - 坐骑交易"),
    ("Core.Good.Paper", "纸张", "文化品", 20, 45, "125%", "造纸坊", "拓展设计 - 文人需求"),
    ("Core.Good.Brush", "毛笔", "文化品", 15, 40, "167%", "笔庄", "拓展设计 - 书法用品"),
    ("Core.Good.Rope", "绳索", "工具", 5, 12, "140%", "麻田", "拓展设计 - 旅行必备"),
]

row = 2
for good in goods_data:
    is_expansion = "拓展设计" in str(good)
    set_row(ws_goods, good[:7], row, is_expansion=is_expansion)
    if len(good) > 7:
        add_expansion_marker(ws_goods, row, 7, good[7])
    row += 1

auto_width(ws_goods)

# ============================================================
# 新增工作表: 旅途事件表 (TravelEvents)
# ============================================================
ws_travel = wb.create_sheet("旅途事件")

travel_headers = ["ID", "名称", "触发条件", "持续时间", "效果", "选项数", "备注"]
set_header(ws_travel, travel_headers)

travel_data = [
    # 天气事件
    ("Core.TravelEvent.Sunny", "晴空万里", "无", "短暂", "旅行速度+20%", 0, "好天气"),
    ("Core.TravelEvent.Rain", "绵绵细雨", "无", "中等", "旅行速度-20%,疲劳+5", 1, "雨天影响"),
    ("Core.TravelEvent.Storm", "暴风雨", "随机", "长", "旅行速度-50%,HP-10,可能中断", 2, "恶劣天气"),
    ("Core.TravelEvent.Snow", "大雪封路", "冬季", "长", "旅行速度-70%,食物消耗+50%", 2, "冬季专属"),
    ("Core.TravelEvent.Fog", "大雾弥漫", "山地", "中", "遭遇战概率+30%,速度-30%", 2, "视野受阻"),
    # 地形事件
    ("Core.TravelEvent.MountainPath", "山路崎岖", "山地", "中等", "速度-30%,可能发现矿石", 2, "地形影响"),
    ("Core.TravelEvent.RiverCrossing", "渡河", "河流", "短", "支付金币或寻找渡口", 2, "交通障碍"),
    ("Core.TravelEvent.ForestPath", "林间小道", "森林", "中", "遭遇野兽或发现草药", 3, "自然环境"),
    ("Core.TravelEvent.Desert", "沙漠戈壁", "沙漠", "长", "水消耗+100%,有绿洲事件", 3, "极端环境"),
    # 拓展旅途事件
    ("Core.TravelEvent.Robbery", "路遇劫匪", "荒野", "战斗", "选择战斗或交钱", 2, "拓展设计 - 抢劫事件"),
    ("Core.TravelEvent.Lost", "迷路", "森林/山地", "长", "消耗额外时间或求助向导", 2, "拓展设计 - 导航挑战"),
    ("Core.TravelEvent.Starfall", "流星雨", "夜间", "短", "许愿获得临时buff", 1, "拓展设计 - 浪漫事件"),
    ("Core.TravelEvent.Plague", "瘟疫蔓延", "城市", "长", "全队HP-20%或花费金币治疗", 2, "拓展设计 - 卫生事件"),
    ("Core.TravelEvent.Festival", "节日庆典", "城市", "短", "商店打折,经验+50%", 2, "拓展设计 - 欢乐事件"),
]

row = 2
for event in travel_data:
    is_expansion = "拓展设计" in str(event)
    set_row(ws_travel, event[:6], row, is_expansion=is_expansion)
    if len(event) > 6:
        add_expansion_marker(ws_travel, row, 6, event[6])
    row += 1

auto_width(ws_travel)

# ============================================================
# 新增工作表: 背包扩容表 (InventoryUpgrades)
# ============================================================
ws_bag = wb.create_sheet("背包扩容")

bag_headers = ["等级", "最大槽位", "最大重量", "解锁条件", "费用", "备注"]
set_header(ws_bag, bag_headers)

bag_data = [
    (1, 20, 50, "初始", "初始", "初始背包"),
    (2, 25, 65, "10级", "500金币", "第一次扩容"),
    (3, 30, 80, "20级", "1500金币", "第二次扩容"),
    (4, 40, 100, "30级", "3000金币", "第三次扩容"),
    (5, 50, 130, "40级", "5000金币", "第四次扩容"),
    (6, 60, 160, "50级", "8000金币", "第五次扩容"),
    (7, 75, 200, "60级", "12000金币", "第六次扩容"),
    (8, 90, 250, "70级", "20000金币", "第七次扩容"),
    # 拓展等级
    (9, 100, 300, "完成7次轮回", "轮回点数:100", "拓展设计 - 永久扩容"),
    (10, 120, 400, "成就:收藏家", "成就解锁", "拓展设计 - 成就奖励"),
]

row = 2
for bag in bag_data:
    is_expansion = "拓展设计" in str(bag)
    set_row(ws_bag, bag[:5], row, is_expansion=is_expansion)
    if len(bag) > 5:
        add_expansion_marker(ws_bag, row, 5, bag[5])
    row += 1

auto_width(ws_bag)

# ============================================================
# 新增工作表: 技能槽位表 (SkillSlots)
# ============================================================
ws_skill_slot = wb.create_sheet("技能槽位")

skill_slot_headers = ["槽位ID", "名称", "技能类型", "数量上限", "解锁条件", "备注"]
set_header(ws_skill_slot, skill_slot_headers)

skill_slot_data = [
    ("Core.SkillSlot.Passive1", "被动技能槽1", "Passive", 1, "初始", "初始解锁"),
    ("Core.SkillSlot.Passive2", "被动技能槽2", "Passive", 1, "15级解锁", "中期解锁"),
    ("Core.SkillSlot.Passive3", "被动技能槽3", "Passive", 1, "35级解锁", "后期解锁"),
    ("Core.SkillSlot.Passive4", "被动技能槽4", "Passive", 1, "完成第一次轮回", "轮回解锁"),
    ("Core.SkillSlot.Active1", "主动技能槽1", "Active", 1, "初始", "初始解锁"),
    ("Core.SkillSlot.Active2", "主动技能槽2", "Active", 1, "25级解锁", "中期解锁"),
    ("Core.SkillSlot.Ultimate", "终极技能槽", "Ultimate", 1, "完成第一次轮回", "轮回解锁"),
    # 拓展槽位
    ("Core.SkillSlot.Passive5", "被动技能槽5", "Passive", 1, "完成5次轮回", "拓展设计 - 轮回解锁"),
    ("Core.SkillSlot.Active3", "主动技能槽3", "Active", 1, "60级解锁", "拓展设计 - 后期解锁"),
    ("Core.SkillSlot.Shared1", "共享技能槽", "Any", 1, "收集10张GR卡", "拓展设计 - 特殊解锁"),
]

row = 2
for slot in skill_slot_data:
    is_expansion = "拓展设计" in str(slot)
    set_row(ws_skill_slot, slot[:5], row, is_expansion=is_expansion)
    if len(slot) > 5:
        add_expansion_marker(ws_skill_slot, row, 5, slot[5])
    row += 1

auto_width(ws_skill_slot)

# ============================================================
# 新增工作表: 抽卡系统 (GachaSystem)
# ============================================================
ws_gacha = wb.create_sheet("抽卡系统")

gacha_headers = ["卡池ID", "名称", "单抽价格", "十连价格", "保底", "保底内容", "备注"]
set_header(ws_gacha, gacha_headers)

gacha_data = [
    ("Core.Gacha.Novice", "新手池", 100, 900, "10抽", "必出R卡", "新手专属"),
    ("Core.Gacha.Standard", "标准池", 300, 2700, "10抽", "必出R卡", "常驻卡池"),
    ("Core.Gacha.Premium", "限定池", 500, 4500, "10抽", "必出SR卡", "限定角色"),
    ("Core.Gacha.Limited", "活动池", 800, 7200, "20抽", "必出SSR卡", "限时活动"),
    ("Core.Gacha.Legendary", "传说池", 1000, 9000, "50抽", "必出UR卡", "顶级卡池"),
    # 拓展卡池
    ("Core.Gacha.Weapon", "武器池", 400, 3600, "15抽", "必出R武器", "拓展设计 - 装备UP"),
    ("Core.Gacha.Skill", "技能池", 400, 3600, "15抽", "必出R技能", "拓展设计 - 技能UP"),
    ("Core.Gacha.Event", "联动池", 600, 5400, "20抽", "必出SR卡", "拓展设计 - 联动活动"),
]

row = 2
for gacha in gacha_data:
    is_expansion = "拓展设计" in str(gacha)
    set_row(ws_gacha, gacha[:6], row, is_expansion=is_expansion)
    if len(gacha) > 6:
        add_expansion_marker(ws_gacha, row, 6, gacha[6])
    row += 1

auto_width(ws_gacha)

# ============================================================
# 新增工作表: 世界地图节点表 (WorldMapNodes)
# ============================================================
ws_map = wb.create_sheet("世界地图节点")

map_headers = ["节点ID", "名称", "类型", "连接节点", "事件列表", "特产", "备注"]
set_header(ws_map, map_headers)

map_data = [
    # 起点区域
    ("Core.Node.Start", "出发点", "Start", "-", "-", "-", "旅程起点"),
    ("Core.Node.Village01", "起始村落", "City", "Start,Wild01", "新手任务,休息恢复", "基础物资", "新手区域"),
    # 荒野区域
    ("Core.Node.Wild01", "东郊荒野", "Wilderness", "Village01,Forest01", "野兽遭遇,随机事件", "草药", "危险区域"),
    ("Core.Node.Wild02", "西山密林", "Wilderness", "Wild01,Dungeon01", "野兽遭遇,宝藏发现", "矿石", "危险区域"),
    # 森林区域
    ("Core.Node.Forest01", "迷雾森林", "Forest", "Wild01,Village02", "迷途,草药采集", "稀有草药", "探索区域"),
    ("Core.Node.Forest02", "古老森林", "Forest", "Forest01,Mountain01", "隐藏事件,精灵", "古董", "传说区域"),
    # 城市区域
    ("Core.Node.Village02", "边境小镇", "City", "Forest01,Market01", "商店,NPC任务", "特产交易", "中转区域"),
    ("Core.Node.City01", "古代王城", "City", "Market01,Market02", "大型商店,竞技场", "奢侈品", "核心城市"),
    # 市场区域
    ("Core.Node.Market01", "东市", "Market", "Village02,City01", "日常交易,随机商人", "生活物资", "商业区域"),
    ("Core.Node.Market02", "西市", "Market", "City01,Mountain01", "珍稀商品,古董商", "古董", "高端市场"),
    # 山地区域
    ("Core.Node.Mountain01", "险峻山路", "Mountain", "Forest02,Market02,Dungeon02", "山贼,矿石", "稀有矿石", "危险区域"),
    ("Core.Node.Mountain02", "云顶之巅", "Mountain", "Mountain01,Boss01", "天气事件,试炼", "天气祝福", "挑战区域"),
    # 副本区域
    ("Core.Node.Dungeon01", "野兽洞穴", "Dungeon", "Wild02,Dungeon02", "连战,宝藏", "兽皮", "初级副本"),
    ("Core.Node.Dungeon02", "废弃矿坑", "Dungeon", "Dungeon01,Mountain01", "连战,高级矿石", "精铁", "中级副本"),
    # BOSS区域
    ("Core.Node.Boss01", "魔王城", "Boss", "Mountain02,Goal", "最终BOSS", "传说装备", "终点前区域"),
    # 终点
    ("Core.Node.Goal", "旅程终点", "Goal", "-", "阶段结算,轮回选择", "-", "阶段结束"),
    # 拓展节点
    ("Core.Node.Temple", "古寺", "Shrine", "Forest02,Village02", "祈福,任务", "神圣祝福", "拓展设计 - 宗教区域"),
    ("Core.Node.Ruins", "遗迹", "Ruins", "Mountain02,Dungeon02", "考古,陷阱", "古代遗物", "拓展设计 - 危险区域"),
    ("Core.Node.Harbor", "港口", "Harbor", "City01,ShipRoute", "船只交易,坐骑", "海产", "拓展设计 - 水路"),
    ("Core.Node.ShipRoute", "航线", "Sea", "Harbor,Goal", "海上事件,海盗", "海外珍宝", "拓展设计 - 水路"),
]

row = 2
for node in map_data:
    is_expansion = "拓展设计" in str(node)
    set_row(ws_map, node[:6], row, is_expansion=is_expansion)
    if len(node) > 6:
        add_expansion_marker(ws_map, row, 6, node[6])
    row += 1

auto_width(ws_map)

# ============================================================
# 新增工作表: 轮回升级表 (PrestigeUpgrades)
# ============================================================
ws_prestige = wb.create_sheet("轮回升级")

prestige_headers = ["升级ID", "名称", "类型", "等级上限", "每次消耗", "效果", "备注"]
set_header(ws_prestige, prestige_headers)

prestige_data = [
    # 基础升级
    ("Core.Prestige.MaxHp", "生命力强化", "Vitality", 10, "10点", "+5%最大HP", "永久加成"),
    ("Core.Prestige.Attack", "攻击力强化", "Attack", 10, "10点", "+5%攻击力", "永久加成"),
    ("Core.Prestige.Defense", "防御力强化", "Defense", 10, "10点", "+5%防御力", "永久加成"),
    ("Core.Prestige.Speed", "速度强化", "Speed", 10, "10点", "+5%速度", "永久加成"),
    ("Core.Prestige.StartingGold", "起始资金", "Economy", 10, "20点", "+100初始金币", "经济加成"),
    ("Core.Prestige.MaxSlot", "背包扩容", "Inventory", 5, "50点", "+10背包槽位", "容量加成"),
    # 特殊升级
    ("Core.Prestige.OfflineRate", "离线收益", "Economy", 5, "30点", "+20%离线收益", "离线加成"),
    ("Core.Prestige.TradeDiscount", "交易折扣", "Social", 5, "25点", "+5%交易折扣", "交易加成"),
    ("Core.Prestige.ExpRate", "经验加成", "Progression", 5, "25点", "+10%经验获取", "成长加成"),
    ("Core.Prestige.CritRate", "暴击率", "Combat", 5, "30点", "+3%暴击率", "战斗加成"),
    # 拓展升级
    ("Core.Prestige.GoldDrop", "金币掉落", "Economy", 5, "40点", "+15%金币掉落", "拓展设计"),
    ("Core.Prestige.ItemFind", "物品发现", "Exploration", 5, "35点", "+10%稀有物品概率", "拓展设计"),
    ("Core.Prestige.SkillSlot", "技能槽位", "Skill", 2, "100点", "+1被动技能槽", "拓展设计 - 稀有升级"),
    ("Core.Prestige.AutoPotion", "自动喝药", "QualityOfLife", 1, "200点", "HP<30%自动使用药水", "拓展设计 - 便利功能"),
]

row = 2
for upgrade in prestige_data:
    is_expansion = "拓展设计" in str(upgrade)
    set_row(ws_prestige, upgrade[:6], row, is_expansion=is_expansion)
    if len(upgrade) > 6:
        add_expansion_marker(ws_prestige, row, 6, upgrade[6])
    row += 1

auto_width(ws_prestige)

# ============================================================
# 新增工作表: 游戏设置表 (GameSettings)
# ============================================================
ws_settings = wb.create_sheet("游戏设置")

settings_headers = ["设置ID", "名称", "默认值", "范围", "说明", "备注"]
set_header(ws_settings, settings_headers)

settings_data = [
    # 基础设置
    ("Core.Setting.MusicVolume", "音乐音量", 0.7, "0-1", "背景音乐音量", "音量控制"),
    ("Core.Setting.SfxVolume", "音效音量", 0.8, "0-1", "音效音量", "音量控制"),
    ("Core.Setting.ShowDamage", "显示伤害", True, "true/false", "战斗中显示伤害数字", "UI设置"),
    ("Core.Setting.AutoSave", "自动存档", True, "true/false", "每5分钟自动存档", "存档设置"),
    ("Core.Setting.SaveInterval", "存档间隔", 300, "秒", "自动存档间隔(秒)", "存档设置"),
    # 战斗设置
    ("Core.Setting.SkipBattleAnimation", "跳过战斗动画", False, "true/false", "快速结束战斗", "战斗设置"),
    ("Core.Setting.AutoEquipment", "自动装备", False, "true/false", "自动穿戴更好装备", "战斗设置"),
    ("Core.Setting.AutoUsePotion", "自动喝药", True, "true/false", "HP低时自动使用药水", "战斗设置"),
    ("Core.Setting.PotionThreshold", "喝药阈值", 0.3, "0-1", "自动喝药的HP百分比", "战斗设置"),
    # 旅行设置
    ("Core.Setting.TravelSpeed", "旅行速度", 1.0, "0.5-3.0", "旅行速度倍率", "旅行设置"),
    ("Core.Setting.ShowTravelTime", "显示旅行时间", True, "true/false", "显示预计到达时间", "旅行设置"),
    # 拓展设置
    ("Core.Setting.ShowDamageFloat", "伤害飘字", True, "true/false", "伤害数字飘动效果", "拓展设计 - UI优化"),
    ("Core.Setting.ConfirmBeforeExit", "退出确认", True, "true/false", "退出前二次确认", "拓展设计 - 误触保护"),
    ("Core.Setting.FastCollectReward", "快速领取", True, "true/false", "离线奖励一键领取", "拓展设计 - 便利功能"),
]

row = 2
for setting in settings_data:
    is_expansion = "拓展设计" in str(setting)
    set_row(ws_settings, setting[:5], row, is_expansion=is_expansion)
    if len(setting) > 5:
        add_expansion_marker(ws_settings, row, 5, setting[5])
    row += 1

auto_width(ws_settings)

# 保存文件
output_path = r"E:\UnityProgram\Game1\Docs\AGENT策划\list.xlsx"
wb.save(output_path)
print(f"Excel文件已更新: {output_path}")
print(f"共 {len(wb.sheetnames)} 个工作表:")
for name in wb.sheetnames:
    print(f"  - {name}")
